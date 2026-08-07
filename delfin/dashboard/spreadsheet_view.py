"""Editable spreadsheet grid for the dashboard file browser.

Renders ``.xlsx``/``.xlsm`` workbooks and delimited text files (``.csv``/``.tsv``)
as an Excel-like HTML grid and writes edits back to disk.

The browser reports edits as an **ordered command journal** rather than a value
diff::

    [{"op": "set", "row": 3, "col": 2, "text": "CCO"},
     {"op": "insert_rows", "at": 5, "count": 1},
     {"op": "delete_cols", "at": 3, "count": 1}]

Both sides replay that journal in the same order, so cell writes and structural
operations stay consistent without either side having to recompute addresses.
Rows and columns are 1-based, matching openpyxl.

Pending edits are pushed to Python as they happen and held there per file, so
selecting another file in the browser cannot silently discard them.

Everything in this module is pure (no ipywidgets, no global state) and openpyxl
is imported lazily, so it can be unit tested in a stripped environment.
"""

from __future__ import annotations

import copy
import csv
from collections import OrderedDict
import datetime as _dt
import decimal as _dec
import html as _html
import io
import json
import os
import re
import shutil
import tempfile
import zipfile
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Dict, List, Mapping, Optional, Sequence, Set, Tuple

# --- Rendering bounds -------------------------------------------------------
# The grid HTML travels over the ipywidgets comm channel on every render, so the
# payload budget binds before the DOM budget does. Large sheets are shown one
# window of rows at a time rather than truncated.
MAX_ROWS = 400
MAX_COLS = 80
MAX_CELLS = 16000
BLANK_TAIL_ROWS = 25
BLANK_TAIL_COLS = 2
MAX_FILE_BYTES = 40 * 1024 * 1024

DEFAULT_COL_WIDTH = 110
MIN_COL_WIDTH = 34
MAX_AUTO_COL_WIDTH = 320
ROW_HEADER_WIDTH = 54

# Reading formulas needs a second parse of the workbook; above this size we skip
# it and edit against the cached values instead.
FORMULA_PASS_MAX_BYTES = 8 * 1024 * 1024

# Scrolling used to re-open the workbook and stream past every row above the
# window -- three times over, once for values, once for the cell formats and
# once for the formulas -- so moving down a 5000-row sheet cost about 0.7 s per
# page, every page, forever.  A sheet that fits this budget is read once in full
# instead and every window after that is a slice of memory.  Above the budget
# the streaming read stands: a sheet of a million rows must not be held.
SHEET_CACHE_MAX_CELLS = 400_000
SHEET_CACHE_ENTRIES = 2

WORKBOOK_SUFFIXES = ('.xlsx', '.xlsm')
DELIMITED_SUFFIXES = ('.csv', '.tsv', '.tab')

# Parts of the xlsx container openpyxl cannot round-trip. Detected with zipfile
# alone so the warning works even where openpyxl is missing.
LOSSY_PARTS = (
    ('xl/charts/', 'charts'),
    ('xl/chartsheets/', 'chart sheets'),
    ('xl/media/', 'embedded images'),
    ('xl/drawings/drawing', 'drawings'),
    ('xl/pivotTables/', 'pivot tables'),
    ('xl/threadedComments/', 'threaded comments'),
    ('xl/slicers', 'slicers'),
    ('customXml/', 'custom XML'),
)

_INT_RE = re.compile(r'^[+-]?(?:0|[1-9][0-9]*)$')
_FLOAT_RE = re.compile(r'^[+-]?(?:(?:0|[1-9][0-9]*)\.[0-9]*|\.[0-9]+)(?:[eE][+-]?[0-9]+)?$')
_SCI_INT_RE = re.compile(r'^[+-]?(?:0|[1-9][0-9]*)[eE][+-]?[0-9]+$')


class SpreadsheetError(RuntimeError):
    """Raised for anything the user should see as a message, not a traceback."""


@dataclass
class SheetData:
    """One window of a worksheet, ready to render."""

    name: str = ''
    values: List[List[str]] = field(default_factory=list)
    formulas: List[List[str]] = field(default_factory=list)
    col_widths: List[int] = field(default_factory=list)
    row_offset: int = 0          # 0-based index of values[0] within the sheet
    total_rows: int = 0
    total_cols: int = 0
    has_formulas: bool = False
    truncated_cols: bool = False
    # Per cell, only what differs from plain: {'b','i','u','bg','fmt'}.
    styles: List[List[Dict[str, Any]]] = field(default_factory=list)

    @property
    def n_rows(self) -> int:
        return len(self.values)

    @property
    def n_cols(self) -> int:
        return len(self.values[0]) if self.values else 0


# ---------------------------------------------------------------------------
# Address helpers
# ---------------------------------------------------------------------------

def col_letter(index0: int) -> str:
    """0 -> 'A', 25 -> 'Z', 26 -> 'AA'."""
    n = int(index0) + 1
    if n < 1:
        raise ValueError(f'column index out of range: {index0}')
    out = ''
    while n > 0:
        n, rem = divmod(n - 1, 26)
        out = chr(ord('A') + rem) + out
    return out


def col_index(letters: str) -> int:
    """'A' -> 0, 'AA' -> 26."""
    text = str(letters or '').strip().upper()
    if not text or not text.isalpha():
        raise ValueError(f'not a column reference: {letters!r}')
    n = 0
    for ch in text:
        n = n * 26 + (ord(ch) - ord('A') + 1)
    return n - 1


# ---------------------------------------------------------------------------
# Value conversion
# ---------------------------------------------------------------------------

def format_cell(value: Any) -> str:
    """Render a stored cell value the way the grid should display it."""
    if value is None:
        return ''
    if isinstance(value, bool):
        return 'TRUE' if value else 'FALSE'
    if isinstance(value, float):
        if value != value or value in (float('inf'), float('-inf')):
            return str(value)
        if value == int(value) and abs(value) < 1e15:
            return str(int(value))
        return repr(value)
    if isinstance(value, _dt.datetime):
        if value.time() == _dt.time(0, 0):
            return value.date().isoformat()
        return value.isoformat(sep=' ')
    if isinstance(value, (_dt.date, _dt.time)):
        return value.isoformat()
    if isinstance(value, _dt.timedelta):
        return str(value)
    return str(value)


# ---------------------------------------------------------------------------
# Cell formatting
# ---------------------------------------------------------------------------

# The formats offered, by the name shown in the menu. Small on purpose: the
# point is a table that reads well, not a format designer. Everything else a
# workbook already carries is left exactly as it is.
NUMBER_FORMATS: Tuple[Tuple[str, str], ...] = (
    ('Standard', 'General'),
    ('1235', '0'),
    ('1234,50', '0.00'),
    ('1.234,50', '#,##0.00'),
    ('1.234,50 €', '#,##0.00\\ "€"'),
    ('12,3 %', '0.0%'),
    ('31.12.2026', 'DD.MM.YYYY'),
    ('Text', '@'),
)
_KNOWN_FORMATS = {code for _label, code in NUMBER_FORMATS}

# Fill colours, light enough that black text stays readable on them. A
# palette rather than a colour picker: the aim is telling rows apart.
FILL_COLOURS: Tuple[Tuple[str, str], ...] = (
    ('None', ''),
    ('Yellow', 'FFF2CC'),
    ('Green', 'E2EFDA'),
    ('Blue', 'DDEBF7'),
    ('Red', 'FCE4EC'),
    ('Orange', 'FDE9D9'),
    ('Purple', 'EDE7F6'),
    ('Grey', 'EDEDED'),
)
_KNOWN_FILLS = {code for _label, code in FILL_COLOURS if code}

_HEX_RE = re.compile(r'^[0-9A-Fa-f]{6}$')


def check_fill(colour: Any) -> str:
    """A fill colour as six hex digits, or '' for none."""
    text = str(colour or '').strip().lstrip('#').upper()
    if not text:
        return ''
    if len(text) == 8:          # openpyxl writes ARGB; the alpha is not ours
        text = text[2:]
    if not _HEX_RE.match(text):
        raise SpreadsheetError(f'{colour!r} is not a colour.')
    return text


def check_number_format(code: Any) -> str:
    """A number format the menu offers, or whatever the cell already had."""
    text = str(code or '').strip()
    if not text:
        return 'General'
    if len(text) > 60:
        raise SpreadsheetError('That number format is too long.')
    return text


def cell_style(cell: Any) -> Dict[str, Any]:
    """The parts of a cell's look that this grid shows and can set.

    Only what differs from plain: an empty dictionary per cell is what
    keeps the markup for a few thousand cells to a readable size.
    """
    style: Dict[str, Any] = {}
    try:
        font = cell.font
        if font is not None:
            if font.bold:
                style['b'] = 1
            if font.italic:
                style['i'] = 1
            if font.underline and font.underline != 'none':
                style['u'] = 1
    except Exception:
        pass
    try:
        fill = cell.fill
        if fill is not None and fill.patternType == 'solid':
            colour = check_fill(getattr(fill.fgColor, 'rgb', '') or '')
            # White is the sheet's own background; marking it would put a
            # colour on every cell of a workbook that has none.
            if colour and colour not in ('FFFFFF', '000000'):
                style['bg'] = colour
    except Exception:
        pass
    try:
        code = str(cell.number_format or 'General')
        if code != 'General':
            style['fmt'] = code
    except Exception:
        pass
    return style


def style_css(style: Mapping[str, Any]) -> str:
    """The style of one cell, as CSS."""
    parts = []
    if style.get('b'):
        # 700, not 600: at this size 600 is barely a difference.
        parts.append('font-weight:700')
    if style.get('i'):
        parts.append('font-style:italic')
    if style.get('u'):
        parts.append('text-decoration:underline')
    background = style.get('bg')
    if background:
        parts.append(f'background:#{background}')
    return ';'.join(parts)


def display_value(value: Any, code: str = 'General') -> str:
    """Show a value the way its number format says to.

    Only the formats this grid offers are worked out here. A workbook can
    carry any format string at all, and one that is not understood shows
    the plain value rather than a guess at what Excel would draw.
    """
    code = str(code or 'General')
    if value is None or code in ('General', '@') or isinstance(value, str):
        return format_cell(value)
    if isinstance(value, (_dt.date, _dt.datetime)) and 'DD' in code.upper():
        return value.strftime('%d.%m.%Y')
    if code not in _KNOWN_FORMATS:
        # A workbook can carry any format string, conditional colours and
        # sections included. Guessing at what Excel would draw would put a
        # number on screen that is not in the file.
        return format_cell(value)
    if not isinstance(value, (int, float)) or isinstance(value, bool):
        return format_cell(value)

    number = float(value)
    if code.endswith('%'):
        decimals = len(code.split('.')[-1].rstrip('%')) if '.' in code else 0
        return _german_number(number * 100, decimals) + ' %'
    decimals = 0
    if '.' in code:
        decimals = len(re.split(r'[.]', code)[-1].split('\\')[0].split('"')[0].strip())
    grouped = ',' in code.split('.')[0]
    text = _german_number(number, decimals, grouped=grouped)
    if '€' in code:
        text += ' €'
    return text


def _german_number(value: float, decimals: int, *, grouped: bool = False) -> str:
    # Half away from zero, which is what a spreadsheet does. Python rounds
    # half to even, so 1234.5 with no decimals would come out as 1234 where
    # Excel writes 1235.
    quantum = _dec.Decimal(1).scaleb(-decimals)
    rounded = _dec.Decimal(repr(value)).quantize(quantum, rounding=_dec.ROUND_HALF_UP)
    text = f'{rounded:,.{decimals}f}' if grouped else f'{rounded:.{decimals}f}'
    # en-US separators out, German ones in, in one pass so the swap cannot
    # trip over its own output.
    return text.translate(str.maketrans({',': '.', '.': ','}))


def coerce_value(text: Any) -> Any:
    """Convert text typed in the grid into the value to store in the cell.

    Deliberately conservative: anything that could be an identifier rather than
    a number stays text. ``007`` is a sample ID, not the integer 7, and NaN/inf
    stay text because they cannot be written to a valid workbook.
    """
    if text is None:
        return None
    raw = str(text)
    if raw == '':
        return None
    if raw.startswith('='):
        return raw  # formula, stored verbatim
    stripped = raw.strip()
    if not stripped:
        return raw
    low = stripped.lower()
    if low == 'true':
        return True
    if low == 'false':
        return False
    if _INT_RE.match(stripped):
        try:
            return int(stripped)
        except ValueError:
            return raw
    if _FLOAT_RE.match(stripped) or _SCI_INT_RE.match(stripped):
        try:
            parsed = float(stripped)
        except ValueError:
            return raw
        if parsed != parsed or parsed in (float('inf'), float('-inf')):
            return raw
        return parsed
    return raw


# ---------------------------------------------------------------------------
# Grid shaping
# ---------------------------------------------------------------------------

def _content_bounds(rows: Sequence[Sequence[str]]) -> Tuple[int, int]:
    """Return (n_rows, n_cols) after dropping fully empty trailing rows/cols."""
    last_row = 0
    last_col = 0
    for r_idx, row in enumerate(rows):
        for c_idx, cell in enumerate(row):
            if cell not in ('', None):
                last_row = r_idx + 1
                if c_idx + 1 > last_col:
                    last_col = c_idx + 1
    return last_row, last_col


def _pad_grid(rows: Sequence[Sequence[str]], n_rows: int, n_cols: int) -> List[List[str]]:
    out: List[List[str]] = []
    for r_idx in range(n_rows):
        row = list(rows[r_idx]) if r_idx < len(rows) else []
        if len(row) < n_cols:
            row.extend([''] * (n_cols - len(row)))
        out.append(row[:n_cols])
    return out


def auto_col_widths(rows: Sequence[Sequence[str]], n_cols: int) -> List[int]:
    """Pick a readable starting width per column from its content."""
    widths = []
    for c_idx in range(n_cols):
        longest = 0
        for row in rows:
            if c_idx < len(row):
                longest = max(longest, len(row[c_idx] or ''))
        px = 8 * min(longest, 40) + 20
        widths.append(max(MIN_COL_WIDTH, min(MAX_AUTO_COL_WIDTH, px or DEFAULT_COL_WIDTH)))
    return widths


def window_rows(total_rows: int, row_offset: int, n_cols: int) -> Tuple[int, int]:
    """Return (start0, per_page) for the row window that fits the cell budget."""
    per_page = MAX_ROWS
    if n_cols > 0:
        per_page = min(per_page, max(25, MAX_CELLS // max(1, n_cols)))
    start = max(0, int(row_offset))
    if total_rows and start >= total_rows:
        start = max(0, ((total_rows - 1) // per_page) * per_page)
    return start, per_page


def _shape_window(
    raw: Sequence[Sequence[str]],
    start0: int,
    per_page: int,
    eff_cols: int,
    max_cols: int,
) -> List[List[str]]:
    """Trim trailing blanks, then leave room to type past the end."""
    n_rows_read, n_cols_read = _content_bounds(raw)
    if start0 == 0:
        n_rows_read = max(n_rows_read, 1)
        n_cols_read = max(n_cols_read, 1)
    n_cols_read = min(max_cols, max(n_cols_read + BLANK_TAIL_COLS, min(eff_cols, n_cols_read)))
    n_rows_view = min(per_page, n_rows_read + BLANK_TAIL_ROWS)
    return _pad_grid(raw, n_rows_view, n_cols_read)


# ---------------------------------------------------------------------------
# Workbook feature inspection (zipfile only -- no openpyxl needed)
# ---------------------------------------------------------------------------

def inspect_workbook_features(path: Path) -> List[str]:
    """Name the parts of a workbook that an openpyxl round trip would drop."""
    found: List[str] = []
    try:
        with zipfile.ZipFile(str(path)) as zf:
            names = zf.namelist()
    except Exception:
        return found
    for prefix, label in LOSSY_PARTS:
        if any(n.startswith(prefix) for n in names):
            found.append(label)
    return found


def describe_lossy_features(features: Sequence[str]) -> str:
    if not features:
        return ''
    listed = ', '.join(features)
    return (
        f'This file contains {listed}. Saving from DELFIN drops them '
        f'(the backup copy keeps them).'
    )


# ---------------------------------------------------------------------------
# Reading
# ---------------------------------------------------------------------------

def _load_workbook(**kwargs):
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - environment dependent
        raise SpreadsheetError(
            'openpyxl is not installed.\n\nInstall it with: pip install openpyxl'
        ) from exc
    return load_workbook(**kwargs)


def list_sheets(path: Path) -> List[str]:
    wb = _load_workbook(filename=str(path), read_only=True, data_only=True)
    try:
        return list(wb.sheetnames)
    finally:
        wb.close()


def _read_window(ws, start0: int, count: int, max_cols: int) -> List[List[str]]:
    rows: List[List[str]] = []
    for raw in ws.iter_rows(
        min_row=start0 + 1,
        max_row=start0 + count,
        min_col=1,
        max_col=max_cols,
        values_only=True,
    ):
        rows.append([format_cell(v) for v in raw])
        if len(rows) >= count:
            break
    return rows


def apply_results(sheet: SheetData, results: 'FormulaResults') -> None:
    """Show what the formulas work out to, in the cells that hold them.

    A workbook that Excel has opened carries its own cached results and
    those are already on screen; one written here has none, and the cell
    would otherwise show =A1*2 where a spreadsheet shows 6. What the cell
    holds is unchanged -- the formula is still there, and pressing F2 or
    double-clicking still shows it, exactly as in Excel.
    """
    if not results or not sheet.formulas:
        return
    for r_idx, row in enumerate(sheet.formulas):
        for c_idx, formula in enumerate(row):
            if not formula:
                continue
            value = results.get(sheet.name,
                                sheet.row_offset + r_idx + 1, c_idx + 1)
            if value is None:
                continue
            shown = format_result(value)
            if r_idx < len(sheet.values) and c_idx < len(sheet.values[r_idx]):
                sheet.values[r_idx][c_idx] = shown


#: (path, mtime_ns, sheet, cols) -> (values, styles, formulas) for the whole
#: sheet.  Small, and dropped as soon as the file changes underneath it.
_SHEET_CACHE: 'OrderedDict[Tuple[str, int, str, int], Tuple[List[List[str]], List[List[Dict[str, Any]]], List[List[str]]]]' = OrderedDict()

#: One shared empty style, so a sheet of unformatted cells does not become a
#: sheet of dictionaries.
_NO_STYLE: Dict[str, Any] = {}


def _sheet_cache_key(path: Path, active: str, cols: int):
    try:
        stamp = path.stat().st_mtime_ns
    except OSError:
        return None
    return (str(path), int(stamp), str(active), int(cols))


def forget_sheet_cache(path: Optional[Path] = None) -> None:
    """Drop what is held for *path*, or everything.  Saving invalidates it."""
    if path is None:
        _SHEET_CACHE.clear()
        return
    target = str(path)
    for key in [k for k in _SHEET_CACHE if k[0] == target]:
        _SHEET_CACHE.pop(key, None)


def _remember_sheet(key, payload) -> None:
    _SHEET_CACHE[key] = payload
    _SHEET_CACHE.move_to_end(key)
    while len(_SHEET_CACHE) > SHEET_CACHE_ENTRIES:
        _SHEET_CACHE.popitem(last=False)


def _read_whole_sheet(path: Path, active: str, total_rows: int, eff_cols: int):
    """Values, formats and formulas for every row, in one pass each.

    Returns None if anything goes wrong, which puts the caller back on the
    streaming read it would have done anyway.
    """
    try:
        book = _load_workbook(filename=str(path), read_only=True, data_only=True)
        try:
            ws = book[active] if active in book.sheetnames else book.worksheets[0]
            values = _pad_grid(
                _read_window(ws, 0, total_rows, eff_cols), total_rows, eff_cols
            )
            styles: List[List[Dict[str, Any]]] = []
            sheet_s = book[active] if active in book.sheetnames else book.worksheets[0]
            for r_idx, row in enumerate(sheet_s.iter_rows(
                    min_row=1, max_row=total_rows, min_col=1, max_col=eff_cols)):
                if r_idx >= total_rows:
                    break
                found: List[Dict[str, Any]] = []
                for c_idx, cell in enumerate(row):
                    style = cell_style(cell)
                    found.append(style or _NO_STYLE)
                    code = style.get('fmt')
                    if (code and cell.value is not None
                            and r_idx < len(values) and c_idx < len(values[r_idx])):
                        values[r_idx][c_idx] = display_value(cell.value, code)
                found.extend(_NO_STYLE for _ in range(eff_cols - len(found)))
                styles.append(found)
            while len(styles) < total_rows:
                styles.append([_NO_STYLE] * eff_cols)
        finally:
            book.close()
    except Exception:
        return None

    formulas: List[List[str]] = []
    try:
        if path.stat().st_size <= FORMULA_PASS_MAX_BYTES:
            book_f = _load_workbook(filename=str(path), read_only=True, data_only=False)
            try:
                ws_f = (book_f[active] if active in book_f.sheetnames
                        else book_f.worksheets[0])
                formulas = _pad_grid(
                    _read_window(ws_f, 0, total_rows, eff_cols), total_rows, eff_cols
                )
            finally:
                book_f.close()
    except Exception:
        formulas = []
    return values, styles, formulas


def sheet_grids_for_formulas(
    path: Path,
    sheet_name: Optional[str] = None,
    *,
    max_cols: int = MAX_COLS,
) -> Optional[Tuple[List[List[str]], List[List[str]]]]:
    """(values, formulas) for a whole sheet, for working formulas out live.

    A formula in the window can refer to rows far above it -- a total under a
    column is the ordinary case -- so the window on screen is not enough to
    work one out.  Returns None for a sheet too large to hold, where the caller
    falls back to the workbook-wide engine.
    """
    path = Path(path)
    try:
        book = _load_workbook(filename=str(path), read_only=True, data_only=True)
    except Exception:
        return None
    try:
        names = list(book.sheetnames)
        if not names:
            return None
        active = sheet_name if sheet_name in names else names[0]
        ws = book[active]
        total_rows = int(ws.max_row or 0)
        total_cols = int(ws.max_column or 0)
    except Exception:
        return None
    finally:
        book.close()
    eff_cols = max(1, min(total_cols or 1, max_cols))
    if not total_rows or total_rows * eff_cols > SHEET_CACHE_MAX_CELLS:
        return None
    key = _sheet_cache_key(path, active, eff_cols)
    payload = _SHEET_CACHE.get(key) if key else None
    if payload is None:
        payload = _read_whole_sheet(path, active, total_rows, eff_cols)
        if payload is None:
            return None
        if key:
            _remember_sheet(key, payload)
    elif key:
        _SHEET_CACHE.move_to_end(key)
    return payload[0], payload[2]


def _slice_rows(grid, start0: int, count: int, cols: int, filler):
    """The window *start0*..*start0+count* of a full-sheet grid."""
    if not grid:
        return []
    out = [list(row[:cols]) for row in grid[start0:start0 + count]]
    for row in out:
        row.extend(filler for _ in range(cols - len(row)))
    return out


def read_xlsx(
    path: Path,
    sheet: Optional[str] = None,
    *,
    row_offset: int = 0,
    max_cols: int = MAX_COLS,
) -> Tuple[List[str], SheetData]:
    """Read one window of one worksheet.

    Returns ``(sheet_names, sheet_data)``. Displayed values come from the cached
    results openpyxl exposes with ``data_only=True``; a second pass collects the
    formulas so editing a formula cell shows the formula, not its result.
    """
    path = Path(path)
    wb = _load_workbook(filename=str(path), read_only=True, data_only=True)
    try:
        names = list(wb.sheetnames)
        if not names:
            raise SpreadsheetError('The workbook has no sheets.')
        active = sheet if sheet in names else names[0]
        ws = wb[active]
        # max_row/max_column come from the sheet's <dimension> record, which some
        # writers omit; treat a missing value as "unknown" rather than zero rows.
        total_rows = int(ws.max_row or 0)
        total_cols = int(ws.max_column or 0)
        truncated_cols = total_cols > max_cols
        eff_cols = max(1, min(total_cols or 1, max_cols))
        start0, per_page = window_rows(total_rows, row_offset, eff_cols)
        # Held only once a second window is asked for: opening a file must stay
        # as quick as it is, and a sheet nobody scrolls is never read in full.
        cache_key = _sheet_cache_key(path, active, eff_cols)
        cached = _SHEET_CACHE.get(cache_key) if cache_key else None
        if (cached is None and cache_key and row_offset > 0
                and 0 < total_rows * eff_cols <= SHEET_CACHE_MAX_CELLS):
            cached = _read_whole_sheet(path, active, total_rows, eff_cols)
            if cached is not None:
                _remember_sheet(cache_key, cached)
        if cached is not None:
            _SHEET_CACHE.move_to_end(cache_key)
            raw_values = _slice_rows(cached[0], start0, per_page, eff_cols, '')
        else:
            raw_values = _read_window(ws, start0, per_page, eff_cols)
    finally:
        wb.close()

    values = _shape_window(raw_values, start0, per_page, eff_cols, max_cols)
    n_rows_view = len(values)
    n_cols_view = len(values[0]) if values else 1

    # What the cells look like, and how their numbers are written. Read from
    # the same window so a cell's value and its format cannot come from
    # different rows.
    styles: List[List[Dict[str, Any]]] = []
    if cached is not None:
        # Already read, formats applied to the values with them.  Both come
        # from the same pass over the same cells, which is the property the
        # streaming version below is careful about too.
        styles = _slice_rows(cached[1], start0, n_rows_view, n_cols_view, _NO_STYLE)
        while len(styles) < n_rows_view:
            styles.append([_NO_STYLE] * n_cols_view)
        for r_idx, row in enumerate(
                _slice_rows(cached[0], start0, n_rows_view, n_cols_view, '')):
            for c_idx, text in enumerate(row):
                if text and r_idx < len(values) and c_idx < len(values[r_idx]):
                    values[r_idx][c_idx] = text
    else:
        try:
            book_s = _load_workbook(filename=str(path), read_only=True, data_only=True)
            try:
                sheet_s = (book_s[active] if active in book_s.sheetnames
                           else book_s.worksheets[0])
                for r_idx, row in enumerate(sheet_s.iter_rows(
                        min_row=start0 + 1, max_row=start0 + n_rows_view,
                        min_col=1, max_col=n_cols_view)):
                    if r_idx >= n_rows_view:
                        break
                    found = []
                    for c_idx, cell in enumerate(row):
                        style = cell_style(cell)
                        found.append(style)
                        # The value is taken from the cell here rather than from
                        # the window of strings: a number and the format that
                        # says how to write it have to come from the same cell.
                        code = style.get('fmt')
                        if (code and cell.value is not None
                                and r_idx < len(values)
                                and c_idx < len(values[r_idx])):
                            values[r_idx][c_idx] = display_value(cell.value, code)
                    found.extend({} for _ in range(n_cols_view - len(found)))
                    styles.append(found)
            finally:
                book_s.close()
            while len(styles) < n_rows_view:
                styles.append([{} for _ in range(n_cols_view)])
        except Exception:
            styles = []

    formulas: List[List[str]] = []
    has_formulas = False
    try:
        if cached is not None:
            raw = _slice_rows(cached[2], start0, n_rows_view, n_cols_view, '')
            raw = _pad_grid(raw, n_rows_view, n_cols_view) if cached[2] else []
        elif path.stat().st_size <= FORMULA_PASS_MAX_BYTES:
            wb_f = _load_workbook(filename=str(path), read_only=True, data_only=False)
            try:
                ws_f = wb_f[active] if active in wb_f.sheetnames else wb_f.worksheets[0]
                raw = _read_window(ws_f, start0, n_rows_view, n_cols_view)
            finally:
                wb_f.close()
            raw = _pad_grid(raw, n_rows_view, n_cols_view)
        else:
            raw = []
        if raw:
            for r_idx, row in enumerate(raw):
                out_row = [cell if cell.startswith('=') else '' for cell in row]
                if any(out_row):
                    has_formulas = True
                    # A workbook never opened by Excel has no cached results, so
                    # show the formula text rather than an empty cell.
                    for c_idx, cell in enumerate(out_row):
                        if cell and not values[r_idx][c_idx]:
                            values[r_idx][c_idx] = cell
                formulas.append(out_row)
    except SpreadsheetError:
        raise
    except Exception:
        formulas = []

    data = SheetData(
        name=active,
        values=values,
        formulas=formulas,
        col_widths=auto_col_widths(values, n_cols_view),
        row_offset=start0,
        total_rows=max(total_rows, start0 + n_rows_view),
        total_cols=total_cols or n_cols_view,
        has_formulas=has_formulas,
        styles=styles,
        truncated_cols=truncated_cols,
    )
    return names, data


@dataclass
class CellHit:
    """A match, addressed as a cell rather than as a character offset."""

    sheet: str
    row: int          # 1-based, as the sheet numbers it
    col: int          # 1-based
    text: str = ''

    @property
    def label(self) -> str:
        # col is 1-based the way the sheet numbers it; col_letter counts
        # from zero.
        return f'{col_letter(self.col - 1)}{self.row}'


# A search that walks the file rather than the loaded window has to stop
# somewhere; where it stopped is reported rather than quietly dropped.
MAX_SEARCH_HITS = 2000


def search_cells(path, term: str, *, delimiter: Optional[str] = None,
                 limit: int = MAX_SEARCH_HITS) -> Tuple[List[CellHit], bool]:
    """Every cell containing ``term``, across the whole file.

    The grid shows one window of rows at a time, so searching what is on
    screen would answer a different question than the user asked. Returns
    the hits and whether the cap stopped it early.

    Reading is streamed: a match on row 90000 must not cost the whole file
    in memory first.
    """
    term = str(term or '')
    if not term:
        return [], False
    needle = term.lower()
    path = Path(path)
    hits: List[CellHit] = []

    def take(sheet_name, row, col, value) -> bool:
        text = '' if value is None else str(value)
        if needle not in text.lower():
            return True
        hits.append(CellHit(sheet=sheet_name, row=row, col=col, text=text))
        return len(hits) < limit

    if path.suffix.lower() in WORKBOOK_SUFFIXES:
        book = _load_workbook(
            filename=str(path), read_only=True, data_only=True)
        try:
            for worksheet in book.worksheets:
                for row_index, row in enumerate(
                        worksheet.iter_rows(values_only=True), start=1):
                    for col_index, value in enumerate(row, start=1):
                        if not take(worksheet.title, row_index, col_index, value):
                            return hits, True
        finally:
            book.close()
        return hits, False

    text, _encoding = decode_delimited(path)
    if delimiter is None:
        delimiter = sniff_delimiter(text, path.suffix)
    reader = csv.reader(io.StringIO(text), delimiter=delimiter)
    for row_index, row in enumerate(reader, start=1):
        for col_index, value in enumerate(row, start=1):
            if not take(path.name, row_index, col_index, value):
                return hits, True
    return hits, False


def window_for_row(row: int, page_rows: int) -> int:
    """The row offset whose window contains ``row`` (1-based).

    Aligned to the paging step the grid already uses, so jumping to a hit
    lands on the same windows the page buttons produce rather than on a
    third set of boundaries.
    """
    step = max(1, int(page_rows))
    return max(0, ((max(1, int(row)) - 1) // step) * step)


# ---------------------------------------------------------------------------
# Text encoding
# ---------------------------------------------------------------------------

def decode_delimited(path: Path) -> Tuple[str, str]:
    """Read a csv/tsv file's text, and say which encoding it was in.

    A CSV exported by a spreadsheet program on a German-language Windows
    system is cp1252, not utf-8. Reading it as utf-8 with errors="replace"
    does not fail -- it turns 'Grünwald' into 'Gr\ufffdnwald', and every
    name and street in a German data set carries one of those bytes, so
    essentially the whole file is quietly damaged.

    The same detection the agent uses, imported rather than repeated: two
    answers to this question would eventually disagree, and the one that
    reads a file has to match the one that writes it back.
    """
    import codecs

    from delfin.agent.office import decode_text

    raw = Path(path).read_bytes()
    text, encoding = decode_text(raw)
    if encoding == 'utf-8-sig' and not raw.startswith(codecs.BOM_UTF8):
        # utf-8-sig decodes a plain utf-8 file happily, but encoding with it
        # writes a byte-order mark -- so reporting it here would put one at
        # the top of every utf-8 file that is saved.
        encoding = 'utf-8'
    return text, encoding


def encode_delimited(text: str, encoding: str) -> bytes:
    """Encode a csv/tsv file's text back into the encoding it came in.

    Writing utf-8 over a cp1252 file would change every umlaut in it,
    including on the rows the edit never touched. Refused with a reason
    when a newly typed character has no place in that encoding, because
    the alternatives are losing the character or silently rewriting the
    whole file in another encoding.
    """
    try:
        return text.encode(encoding)
    except (UnicodeEncodeError, LookupError) as exc:
        bad = getattr(exc, 'object', '')[
            getattr(exc, 'start', 0):getattr(exc, 'end', 0)] or '?'
        raise SpreadsheetError(
            f'This file is {encoding}, and {bad!r} cannot be written in it. '
            'Save the file as UTF-8 once, then this character can be used.'
        ) from exc


# ---------------------------------------------------------------------------
# What a formula works out to
# ---------------------------------------------------------------------------

# Evaluating a workbook builds a graph of every cell in it, which costs
# seconds on a large one. Bounded so that opening a file never turns into a
# wait, and the bound is reported rather than passed off as "no result".
MAX_FORMULA_CELLS = 3000
MAX_FORMULA_BYTES = 8 * 1024 * 1024

_CELL_KEY_RE = re.compile(r"^'\[[^\]]*\](?P<sheet>[^']*)'!(?P<col>[A-Z]{1,3})(?P<row>\d+)$")


class FormulaResults:
    """What a workbook's formulas work out to, per sheet and cell.

    ``note`` says why a value is missing when one is: the engine is not
    installed, the workbook was too large to evaluate, or that particular
    formula is beyond it. A cell with no result keeps showing its formula,
    which is honest -- a wrong number would not be.
    """

    def __init__(self, values: Optional[Dict[str, Dict[Tuple[int, int], Any]]] = None,
                 note: str = ''):
        self.values = values or {}
        self.note = note

    def get(self, sheet: str, row: int, col: int) -> Any:
        return self.values.get(str(sheet).upper(), {}).get((int(row), int(col)))

    def __bool__(self) -> bool:
        return bool(self.values)


def count_formulas(path: Path) -> int:
    """How many formula cells a workbook holds, without evaluating any."""
    book = _load_workbook(filename=str(path), read_only=True, data_only=False)
    try:
        total = 0
        for worksheet in book.worksheets:
            for row in worksheet.iter_rows():
                for cell in row:
                    if isinstance(cell.value, str) and cell.value.startswith('='):
                        total += 1
                        if total > MAX_FORMULA_CELLS:
                            return total
        return total
    finally:
        book.close()


def evaluate_workbook(path: Path) -> FormulaResults:
    """Work out every formula in a workbook.

    Runs on the file, so it shows what the file says -- unsaved edits are
    not in it yet, which is why the caller re-runs this after a save.
    """
    path = Path(path)
    try:
        size = path.stat().st_size
    except OSError as exc:
        return FormulaResults(note=f'could not read the file: {exc}')
    if size > MAX_FORMULA_BYTES:
        return FormulaResults(
            note=f'workbook is {size / (1024 * 1024):.0f} MB; formulas are '
                 'shown as written rather than worked out')
    try:
        import formulas
    except ImportError:
        return FormulaResults(
            note='formula results need the "formulas" package '
                 '(pip install formulas)')

    try:
        if count_formulas(path) > MAX_FORMULA_CELLS:
            return FormulaResults(
                note=f'more than {MAX_FORMULA_CELLS} formulas; they are shown '
                     'as written rather than worked out')
    except Exception:
        pass

    try:
        # The engine draws a progress bar. In a notebook that is not a
        # terminal it lands in the dashboard's own output, under the grid.
        import contextlib
        import io

        sink = io.StringIO()
        with contextlib.redirect_stdout(sink), contextlib.redirect_stderr(sink):
            model = formulas.ExcelModel().loads(str(path)).finish()
            solution = model.calculate()
    except Exception as exc:  # noqa: BLE001 - the engine raises many kinds
        return FormulaResults(note=f'could not work out the formulas: {exc}')

    values: Dict[str, Dict[Tuple[int, int], Any]] = {}
    for key, value in (solution or {}).items():
        match = _CELL_KEY_RE.match(str(key))
        if match is None:
            continue          # ranges and named things, not single cells
        cell = _single_value(value)
        if cell is None:
            continue
        sheet = match.group('sheet').upper()
        row = int(match.group('row'))
        col = col_index(match.group('col')) + 1
        values.setdefault(sheet, {})[(row, col)] = cell
    return FormulaResults(values)


def _single_value(value: Any) -> Any:
    """Unwrap the engine's 1x1 matrix. Anything else is not a cell value."""
    try:
        array = getattr(value, 'value', value)
        while hasattr(array, '__len__') and not isinstance(array, (str, bytes)):
            if len(array) != 1:
                return None
            array = array[0]
        return array
    except Exception:
        return None


def format_result(value: Any) -> str:
    """Write a worked-out value the way a spreadsheet shows it."""
    if value is None:
        return ''
    if isinstance(value, bool):
        return 'TRUE' if value else 'FALSE'
    if isinstance(value, float):
        if value != value or value in (float('inf'), float('-inf')):
            return '#NUM!'
        if float(value).is_integer():
            return str(int(value))
        return f'{value:.10g}'
    return str(value)


# ---------------------------------------------------------------------------
# Filling by dragging the handle
# ---------------------------------------------------------------------------

# An A1-style reference, with the two places a $ can lock it.
_REF_RE = re.compile(r'(\$?)([A-Za-z]{1,3})(\$?)([1-9][0-9]{0,6})')
# Text ending in digits: "Pos 7" continues as "Pos 8".
_TRAILING_NUMBER_RE = re.compile(r'^(.*?)(\d+)$')
_ISO_DATE_RE = re.compile(r'^(\d{4})-(\d{2})-(\d{2})$')


def shift_formula(text: str, d_row: int, d_col: int) -> str:
    """Move a formula's relative references by (d_row, d_col).

    What makes dragging a formula down useful: =B2*C2 has to become
    =B3*C3, or the copies all read the row they came from. A reference
    locked with $ stays put, which is the whole reason $ exists.

    Text inside quotes is left alone -- ="A1 bis A9" is a label, not two
    references.
    """
    text = str(text)
    if not text.startswith('='):
        return text

    out = []
    index = 0
    in_string = False
    while index < len(text):
        char = text[index]
        if char == '"':
            in_string = not in_string
            out.append(char)
            index += 1
            continue
        if in_string:
            out.append(char)
            index += 1
            continue
        match = _REF_RE.match(text, index)
        if match is None:
            out.append(char)
            index += 1
            continue
        col_lock, col_text, row_lock, row_text = match.groups()
        column = col_index(col_text.upper()) + 1
        row = int(row_text)
        if not col_lock:
            column += d_col
        if not row_lock:
            row += d_row
        if column < 1 or row < 1:
            # Off the sheet: Excel writes #REF! and so do we, rather than
            # silently clamping to A1 and reading the wrong cell.
            out.append('#REF!')
        else:
            out.append(f'{col_lock}{col_letter(column - 1)}{row_lock}{row}')
        index = match.end()
    return ''.join(out)


def _as_number(text: str) -> Optional[float]:
    try:
        return float(str(text).strip().replace(',', '.'))
    except (TypeError, ValueError):
        return None


def _format_number(value: float, like: str) -> str:
    """Write a continued number the way the ones it continues are written."""
    text = str(like).strip()
    decimals = 0
    if '.' in text or ',' in text:
        decimals = len(re.split(r'[.,]', text)[-1])
    formatted = f'{value:.{decimals}f}' if decimals else f'{round(value):d}'
    return formatted.replace('.', ',') if ',' in text else formatted


def _as_date(text: str):
    match = _ISO_DATE_RE.match(str(text).strip())
    if match is None:
        return None
    try:
        return _dt.date(*(int(part) for part in match.groups()))
    except ValueError:
        return None


def fill_series(source: Sequence[str], count: int, *,
                d_row: int = 0, d_col: int = 0) -> List[str]:
    """Continue ``source`` for ``count`` more cells.

    The rules are the ones a spreadsheet user expects, and no more:

    * formulas move with the cell, so a column of =B2*C2 keeps pointing at
      its own row;
    * two or more numbers set a step and it carries on -- one number on its
      own is copied, which is what Excel does and what stops a column of
      identical prices turning into a count;
    * a single date steps by a day, several set their own step;
    * text ending in digits counts on ("Pos 7", "Pos 8");
    * anything else repeats, cycling through the block that was dragged.
    """
    source = [('' if v is None else str(v)) for v in source]
    count = max(0, int(count))
    if not source or not count:
        return []

    if all(v.startswith('=') for v in source if v.strip()) and any(
            v.startswith('=') for v in source):
        return [
            shift_formula(source[i % len(source)],
                          d_row * (i // len(source) + 1),
                          d_col * (i // len(source) + 1))
            for i in range(count)
        ]

    numbers = [_as_number(v) for v in source]
    if all(n is not None for n in numbers) and len(numbers) >= 2:
        step = numbers[-1] - numbers[-2]
        return [_format_number(numbers[-1] + step * (i + 1), source[-1])
                for i in range(count)]

    dates = [_as_date(v) for v in source]
    if all(d is not None for d in dates):
        step = ((dates[-1] - dates[-2]).days if len(dates) >= 2 else 1) or 1
        return [(dates[-1] + _dt.timedelta(days=step * (i + 1))).isoformat()
                for i in range(count)]

    if len(source) == 1 and numbers[0] is None:
        match = _TRAILING_NUMBER_RE.match(source[0])
        if match is not None:
            head, digits = match.groups()
            width = len(digits) if digits.startswith('0') else 0
            return [f'{head}{str(int(digits) + i + 1).zfill(width)}'
                    for i in range(count)]

    return [source[i % len(source)] for i in range(count)]


def fill_block(source: Sequence[Sequence[str]], rows: int, cols: int
               ) -> List[List[str]]:
    """Extend a rectangular block downwards or to the right.

    One direction at a time, the way the handle is dragged. ``rows`` and
    ``cols`` are how many cells to add beyond the block itself.
    """
    grid = [list(row) for row in source]
    if not grid or not grid[0]:
        return []
    rows, cols = max(0, int(rows)), max(0, int(cols))
    if rows and cols:
        raise SpreadsheetError('A fill goes down or across, not both at once.')

    if rows:
        height = len(grid)
        columns = list(zip(*grid))
        filled = [fill_series(list(column), rows, d_row=height)
                  for column in columns]
        return [list(values) for values in zip(*filled)]

    width = len(grid[0])
    return [fill_series(row, cols, d_col=width) for row in grid]


def sniff_delimiter(sample: str, suffix: str = '') -> str:
    """Pick the field separator for a delimited text file."""
    if str(suffix).lower() in ('.tsv', '.tab'):
        return '\t'
    head = '\n'.join(sample.splitlines()[:20])
    if not head.strip():
        return ','
    try:
        return csv.Sniffer().sniff(head, delimiters=',;\t|').delimiter
    except Exception:
        counts = {d: head.count(d) for d in (';', ',', '\t', '|')}
        best = max(counts, key=lambda d: counts[d])
        return best if counts[best] else ','


def read_delimited(
    path: Path,
    *,
    row_offset: int = 0,
    max_cols: int = MAX_COLS,
    delimiter: Optional[str] = None,
) -> Tuple[SheetData, str]:
    """Read a csv/tsv file into the same grid shape as a worksheet."""
    path = Path(path)
    text, _encoding = decode_delimited(path)
    if delimiter is None:
        delimiter = sniff_delimiter(text, path.suffix)
    all_rows = [list(r) for r in csv.reader(io.StringIO(text), delimiter=delimiter)]
    total_rows = len(all_rows)
    total_cols = max((len(r) for r in all_rows), default=1)
    truncated_cols = total_cols > max_cols
    eff_cols = max(1, min(total_cols or 1, max_cols))
    start0, per_page = window_rows(total_rows, row_offset, eff_cols)
    window = [[('' if c is None else str(c)) for c in r[:eff_cols]]
              for r in all_rows[start0:start0 + per_page]]
    values = _shape_window(window, start0, per_page, eff_cols, max_cols)

    data = SheetData(
        name=path.name,
        values=values,
        formulas=[],
        col_widths=auto_col_widths(values, len(values[0]) if values else 1),
        row_offset=start0,
        total_rows=max(total_rows, start0 + len(values)),
        total_cols=total_cols or 1,
        has_formulas=False,
        truncated_cols=truncated_cols,
    )
    return data, delimiter


def grid_to_tsv(sheet: SheetData) -> str:
    """Plain-text fallback used by the tab's search and copy buttons."""
    return '\n'.join('\t'.join(row) for row in sheet.values)


# ---------------------------------------------------------------------------
# Edit journal
# ---------------------------------------------------------------------------

def validate_ops(ops: Any) -> List[Dict[str, Any]]:
    """Sanity-check an edit journal that arrived from the browser."""
    if not isinstance(ops, (list, tuple)):
        raise ValueError('ops must be a list')
    clean: List[Dict[str, Any]] = []
    for op in ops:
        if not isinstance(op, dict):
            raise ValueError(f'bad op: {op!r}')
        kind = op.get('op')
        if kind == 'set':
            row = int(op.get('row', 0))
            col = int(op.get('col', 0))
            if row < 1 or col < 1 or row > 1048576 or col > 16384:
                raise ValueError(f'cell out of range: r{row}c{col}')
            text = op.get('text', '')
            clean.append({'op': 'set', 'row': row, 'col': col, 'text': '' if text is None else str(text)})
        elif kind == 'format':
            box = {}
            for key in ('r1', 'c1', 'r2', 'c2'):
                box[key] = int(op.get(key, 0))
            if (box['r1'] < 1 or box['c1'] < 1 or box['r2'] < box['r1']
                    or box['c2'] < box['c1']
                    or box['r2'] > 1048576 or box['c2'] > 16384):
                raise ValueError(f'bad range: {box}')
            clean_op: Dict[str, Any] = {'op': 'format', **box}
            for key in ('bold', 'italic', 'underline'):
                if op.get(key) is not None:
                    clean_op[key] = bool(op.get(key))
            if op.get('fill') is not None:
                clean_op['fill'] = check_fill(op.get('fill'))
            if op.get('number_format') is not None:
                clean_op['number_format'] = check_number_format(
                    op.get('number_format'))
            if len(clean_op) == 5:
                raise ValueError('a format op has to change something')
            clean.append(clean_op)
        elif kind in ('insert_rows', 'delete_rows', 'insert_cols', 'delete_cols'):
            at = int(op.get('at', 0))
            count = int(op.get('count', 1))
            if at < 1 or count < 1 or count > 1000:
                raise ValueError(f'bad {kind}: at={at} count={count}')
            clean.append({'op': kind, 'at': at, 'count': count})
        else:
            raise ValueError(f'unknown op: {kind!r}')
    return clean


def replay_ops(sheet: SheetData, ops: Sequence[Mapping[str, Any]]) -> Set[Tuple[int, int]]:
    """Apply a journal to an in-memory window so pending edits stay visible.

    Mutates ``sheet`` in place and returns the window coordinates (row index,
    column index, both 0-based) that should be marked as unsaved.
    """
    dirty: Set[Tuple[int, int]] = set()

    def _width() -> int:
        return len(sheet.values[0]) if sheet.values else 0

    def _restyle(op: Mapping[str, Any]) -> None:
        for row in range(op['r1'], op['r2'] + 1):
            r_idx = row - 1 - sheet.row_offset
            if not 0 <= r_idx < len(sheet.values):
                continue
            while len(sheet.styles) <= r_idx:
                sheet.styles.append([{} for _ in range(_width())])
            row_styles = sheet.styles[r_idx]
            while len(row_styles) < _width():
                row_styles.append({})
            for col in range(op['c1'], op['c2'] + 1):
                c_idx = col - 1
                if not 0 <= c_idx < len(row_styles):
                    continue
                style = dict(row_styles[c_idx] or {})
                for key, mark in (('bold', 'b'), ('italic', 'i'),
                                  ('underline', 'u')):
                    if key in op:
                        if op[key]:
                            style[mark] = 1
                        else:
                            style.pop(mark, None)
                if 'fill' in op:
                    if op['fill']:
                        style['bg'] = op['fill']
                    else:
                        style.pop('bg', None)
                if 'number_format' in op:
                    code = op['number_format']
                    if code and code != 'General':
                        style['fmt'] = code
                    else:
                        style.pop('fmt', None)
                row_styles[c_idx] = style
                dirty.add((r_idx, c_idx))

    for op in ops:
        kind = op['op']
        if kind == 'format':
            _restyle(op)
            continue
        if kind == 'set':
            r = op['row'] - 1 - sheet.row_offset
            c = op['col'] - 1
            if r < 0 or c < 0 or r >= len(sheet.values):
                continue
            while c >= _width():
                for row in sheet.values:
                    row.append('')
                for row in sheet.formulas:
                    row.append('')
            text = op['text']
            sheet.values[r][c] = text
            if sheet.formulas and r < len(sheet.formulas):
                sheet.formulas[r][c] = text if text.startswith('=') else ''
            dirty.add((r, c))
        elif kind == 'insert_rows':
            r = op['at'] - 1 - sheet.row_offset
            if r < 0 or r > len(sheet.values):
                continue
            for _ in range(op['count']):
                sheet.values.insert(r, [''] * _width())
                if sheet.formulas:
                    sheet.formulas.insert(r, [''] * _width())
            dirty = {(rr + op['count'] if rr >= r else rr, cc) for rr, cc in dirty}
        elif kind == 'delete_rows':
            r = op['at'] - 1 - sheet.row_offset
            if r < 0 or r >= len(sheet.values):
                continue
            n = min(op['count'], len(sheet.values) - r)
            del sheet.values[r:r + n]
            if sheet.formulas:
                del sheet.formulas[r:r + n]
            dirty = {(rr - n if rr >= r + n else rr, cc) for rr, cc in dirty if not (r <= rr < r + n)}
        elif kind == 'insert_cols':
            c = op['at'] - 1
            if c < 0 or c > _width():
                continue
            for _ in range(op['count']):
                for row in sheet.values:
                    row.insert(c, '')
                for row in sheet.formulas:
                    row.insert(c, '')
                sheet.col_widths.insert(min(c, len(sheet.col_widths)), DEFAULT_COL_WIDTH)
            dirty = {(rr, cc + op['count'] if cc >= c else cc) for rr, cc in dirty}
        elif kind == 'delete_cols':
            c = op['at'] - 1
            if c < 0 or c >= _width():
                continue
            n = min(op['count'], _width() - c)
            for row in sheet.values:
                del row[c:c + n]
            for row in sheet.formulas:
                del row[c:c + n]
            del sheet.col_widths[c:c + n]
            dirty = {(rr, cc - n if cc >= c + n else cc) for rr, cc in dirty if not (c <= cc < c + n)}

    return dirty


# ---------------------------------------------------------------------------
# Writing
# ---------------------------------------------------------------------------

def backup_path_for(path: Path) -> Path:
    path = Path(path)
    return path.with_name(f'{path.stem}.bak{path.suffix}')


def make_backup(path: Path, *, folder: Optional[Path] = None,
                versioned: bool = False) -> Optional[Path]:
    """Copy the file aside before it is changed. Returns the copy, or None.

    Two shapes, because two places look for the copy. The calculations
    browser keeps one ``name.bak.ext`` beside the file, which is where its
    users have always found it, and takes it only once. Office keeps a
    numbered history -- that part is :mod:`delfin.doc_backup`, shared with
    the agent so both fill one sequence instead of writing two histories of
    the same file into the same folder.
    """
    path = Path(path)
    if versioned:
        from delfin import doc_backup

        return doc_backup.make_backup(path, folder=folder)

    target = backup_path_for(path)
    if folder is not None:
        folder = Path(folder)
        try:
            folder.mkdir(parents=True, exist_ok=True)
        except OSError:
            pass          # unwritable: beside the file is better than nowhere
        else:
            target = folder / target.name
    if target.exists():
        return None
    shutil.copy2(str(path), str(target))
    return target


def _atomic_replace(tmp_path: str, target: Path) -> None:
    os.replace(tmp_path, str(target))


def write_text_atomic(path: Path, text: str, *, newline: str = '') -> None:
    """Write text through a temporary file in the same directory, then swap.

    A failure part way through leaves the original file untouched instead of
    truncated.
    """
    path = Path(path)
    tmp_fd, tmp_name = tempfile.mkstemp(dir=str(path.parent), prefix='.dtext-', suffix=path.suffix)
    try:
        with os.fdopen(tmp_fd, 'w', encoding='utf-8', newline=newline) as fh:
            fh.write(text)
        _atomic_replace(tmp_name, path)
    except BaseException:
        try:
            os.unlink(tmp_name)
        except OSError:
            pass
        raise


MAX_SHEET_NAME = 31
# Characters Excel refuses in a sheet name. Rejected here rather than
# silently rewritten: a sheet the user cannot find by the name they typed is
# worse than being told the name will not do.
_BAD_SHEET_CHARS = set(r'[]:*?/\\')


def check_sheet_name(name: str, existing: Sequence[str] = ()) -> str:
    """Return a usable sheet name, or say why this one is not."""
    name = str(name or '').strip()
    if not name:
        raise SpreadsheetError('A sheet needs a name.')
    if len(name) > MAX_SHEET_NAME:
        raise SpreadsheetError(
            f'A sheet name can be at most {MAX_SHEET_NAME} characters.')
    bad = sorted(set(name) & _BAD_SHEET_CHARS)
    if bad:
        raise SpreadsheetError(
            'A sheet name cannot contain ' + ' '.join(repr(c) for c in bad) + '.')
    if name.startswith("'") or name.endswith("'"):
        raise SpreadsheetError("A sheet name cannot start or end with '.")
    lowered = {str(e).lower() for e in existing}
    if name.lower() in lowered:
        raise SpreadsheetError(f'There is already a sheet called {name!r}.')
    return name


def _sort_key(value: Any):
    """Order the way a spreadsheet orders: numbers, then text, blanks last.

    Blanks sink to the bottom whichever way round the sort goes, which is
    what Excel does -- a column sorted descending should not start with its
    empty rows.
    """
    if value is None or (isinstance(value, str) and not value.strip()):
        return (2, 0.0, '')
    if isinstance(value, bool):
        return (1, 0.0, 'TRUE' if value else 'FALSE')
    if isinstance(value, (int, float)):
        return (0, float(value), '')
    if isinstance(value, (_dt.date, _dt.datetime)):
        return (0, float(_dt.datetime.combine(
            value if isinstance(value, _dt.date) and not isinstance(
                value, _dt.datetime) else value.date(),
            _dt.time()).timestamp()), '')
    text = str(value)
    number = _as_number(text)
    if number is not None:
        return (0, number, '')
    return (1, 0.0, text.casefold())


def looks_like_a_header(rows: Sequence[Sequence[Any]]) -> bool:
    """Whether the first row is a heading rather than data.

    Read across the whole row, not down the column being sorted: a column
    of names under the heading "Name" is text over text and says nothing,
    while the "Betrag" beside it sits over numbers and settles it.

    With nothing to go on -- text everywhere -- it is taken as a heading,
    which is what a spreadsheet assumes too. Sorting a heading into the
    middle of the data is the worse mistake of the two, and not one that
    can be spotted by looking at the top of the sheet.
    """
    if len(rows) < 2:
        return False
    first = rows[0]
    if not any(isinstance(cell, str) and cell.strip() for cell in first):
        return False
    for row in rows[1:6]:
        for index, top in enumerate(first):
            if index >= len(row):
                continue
            below = row[index]
            if below is None or not isinstance(top, str) or not top.strip():
                continue
            if _as_number(str(top)) is not None:
                continue          # the heading is itself a number
            if not isinstance(below, str) or _as_number(str(below)) is not None:
                return True
    return True


def sort_sheet(path: Path, sheet_name: str, column: int, *,
               descending: bool = False, header: Optional[bool] = None,
               backup: bool = True, backup_dir: Optional[Path] = None
               ) -> Tuple[Optional[Path], bool]:
    """Sort a whole sheet by one column, in the file. Returns (backup, header).

    The whole sheet, not the window on screen: sorting the rows that happen
    to be visible would interleave them with the ones that are not, and the
    file would be scrambled in a way nobody could see.

    Formulas move with their row. A reference that pointed at the row it
    sits in still does afterwards -- which is the common case and the one
    that silently reads the wrong cell if it is left alone.
    """
    path = Path(path)
    column = int(column)
    if column < 1:
        raise SpreadsheetError('There is no such column to sort by.')

    book = _load_workbook(filename=str(path), keep_vba=path.suffix.lower() == '.xlsm')
    try:
        if sheet_name not in book.sheetnames:
            raise SpreadsheetError(f'There is no sheet called {sheet_name!r}.')
        worksheet = book[sheet_name]
        rows = list(worksheet.iter_rows())
        if len(rows) < 2:
            return None, False
        values = [[cell.value for cell in row] for row in rows]
        if column > len(values[0]) and all(column > len(v) for v in values):
            raise SpreadsheetError('There is no such column to sort by.')

        keep = looks_like_a_header(values) if header is None else bool(header)
        start = 1 if keep else 0
        index = column - 1

        def _value(r):
            return values[r][index] if index < len(values[r]) else None

        def _blank(r):
            cell = _value(r)
            return cell is None or (isinstance(cell, str) and not cell.strip())

        # Blanks last whichever way the sort goes: reversing the order would
        # otherwise start a descending column with its empty rows. Sorting
        # them separately is also what keeps them in the order they were in.
        filled = [r for r in range(start, len(rows)) if not _blank(r)]
        empty = [r for r in range(start, len(rows)) if _blank(r)]
        filled.sort(key=lambda r: _sort_key(_value(r)), reverse=bool(descending))
        body = filled + empty
        if body == list(range(start, len(rows))):
            return None, keep

        made = (make_backup(path, folder=backup_dir, versioned=bool(backup_dir))
                if backup else None)

        # Read everything out before writing anything back: the source and
        # the target of a move overlap.
        taken = []
        for source in body:
            row = []
            for cell in rows[source]:
                row.append((cell.value, copy.copy(cell._style)))
            taken.append((source, row))

        for offset, (source, row) in enumerate(taken):
            target = start + offset
            shift = target - source
            for col_index, (value, style) in enumerate(row, start=1):
                cell = worksheet.cell(row=target + 1, column=col_index)
                if isinstance(value, str) and value.startswith('=') and shift:
                    value = shift_formula(value, shift, 0)
                cell.value = value
                cell._style = style
        _save_workbook(book, path)
    finally:
        book.close()
    return made, keep


def add_sheet(path: Path, name: str, *, backup: bool = True,
              backup_dir: Optional[Path] = None) -> Tuple[Optional[Path], str]:
    """Add an empty sheet to a workbook. Returns (backup, the name used)."""
    path = Path(path)
    book = _load_workbook(filename=str(path), keep_vba=path.suffix.lower() == '.xlsm')
    try:
        name = check_sheet_name(name, book.sheetnames)
        book.create_sheet(title=name)
        made = (make_backup(path, folder=backup_dir, versioned=bool(backup_dir))
                if backup else None)
        _save_workbook(book, path)
    finally:
        book.close()
    return made, name


def rename_sheet(path: Path, old: str, new: str, *, backup: bool = True,
                 backup_dir: Optional[Path] = None) -> Tuple[Optional[Path], str]:
    """Rename one sheet of a workbook."""
    path = Path(path)
    book = _load_workbook(filename=str(path), keep_vba=path.suffix.lower() == '.xlsm')
    try:
        if old not in book.sheetnames:
            raise SpreadsheetError(f'There is no sheet called {old!r}.')
        if str(new).strip() == old:
            return None, old
        others = [n for n in book.sheetnames if n != old]
        new = check_sheet_name(new, others)
        made = (make_backup(path, folder=backup_dir, versioned=bool(backup_dir))
                if backup else None)
        book[old].title = new
        _save_workbook(book, path)
    finally:
        book.close()
    return made, new


def drop_sheet(path: Path, name: str, *, backup: bool = True,
               backup_dir: Optional[Path] = None) -> Tuple[Optional[Path], str]:
    """Remove a sheet. Returns (backup, the sheet to show instead)."""
    path = Path(path)
    book = _load_workbook(filename=str(path), keep_vba=path.suffix.lower() == '.xlsm')
    try:
        if name not in book.sheetnames:
            raise SpreadsheetError(f'There is no sheet called {name!r}.')
        if len(book.sheetnames) <= 1:
            raise SpreadsheetError(
                'This is the only sheet in the workbook; a workbook has to '
                'keep one.')
        index = book.sheetnames.index(name)
        made = (make_backup(path, folder=backup_dir, versioned=bool(backup_dir))
                if backup else None)
        del book[name]
        remaining = book.sheetnames
        _save_workbook(book, path)
    finally:
        book.close()
    return made, remaining[min(index, len(remaining) - 1)]


def _save_workbook(book: Any, path: Path) -> None:
    """Write a workbook over itself without risking a half-written file."""
    tmp_fd, tmp_name = tempfile.mkstemp(
        dir=str(path.parent), prefix='.dsheet-', suffix=path.suffix)
    os.close(tmp_fd)
    try:
        book.save(tmp_name)
        _atomic_replace(tmp_name, path)
    except BaseException:
        try:
            os.unlink(tmp_name)
        except OSError:
            pass
        raise


def _apply_format(worksheet: Any, op: Mapping[str, Any]) -> None:
    """Write one format op into a worksheet.

    Each attribute is set on top of what the cell already has: a workbook
    carries fonts, borders and colours this grid never shows, and replacing
    the whole style would throw those away for the sake of one bold.
    """
    from openpyxl.styles import Font, PatternFill

    for row in range(op['r1'], op['r2'] + 1):
        for col in range(op['c1'], op['c2'] + 1):
            cell = worksheet.cell(row=row, column=col)
            if any(key in op for key in ('bold', 'italic', 'underline')):
                font = cell.font
                cell.font = Font(
                    name=font.name, size=font.size, color=font.color,
                    bold=op.get('bold', font.bold),
                    italic=op.get('italic', font.italic),
                    underline=('single' if op['underline'] else None)
                    if 'underline' in op else font.underline,
                    strike=font.strike, vertAlign=font.vertAlign,
                )
            if 'fill' in op:
                colour = op['fill']
                cell.fill = (PatternFill('solid', fgColor=f'FF{colour}')
                             if colour else PatternFill(fill_type=None))
            if 'number_format' in op:
                cell.number_format = op['number_format'] or 'General'


def apply_ops_xlsx(
    path: Path,
    sheet_name: str,
    ops: Any,
    *,
    backup: bool = True,
    backup_dir: Optional[Path] = None,
) -> Optional[Path]:
    """Replay an edit journal against a workbook and save it in place.

    Loads *without* ``data_only`` -- loading with it and saving would replace
    every formula in the workbook with its cached value. Writes to a temporary
    file in the same directory and swaps it in atomically, so a failure part way
    through cannot leave a truncated workbook behind.

    Returns the backup path if one was created.
    """
    path = Path(path)
    clean = validate_ops(ops)
    if not clean:
        return None

    kwargs: Dict[str, Any] = {'filename': str(path), 'data_only': False}
    if path.suffix.lower() == '.xlsm':
        kwargs['keep_vba'] = True
    wb = _load_workbook(**kwargs)
    made = (make_backup(path, folder=backup_dir, versioned=bool(backup_dir))
            if backup else None)
    tmp_fd, tmp_name = tempfile.mkstemp(dir=str(path.parent), prefix='.dsheet-', suffix=path.suffix)
    os.close(tmp_fd)
    try:
        ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.worksheets[0]
        for op in clean:
            kind = op['op']
            if kind == 'set':
                ws.cell(row=op['row'], column=op['col']).value = coerce_value(op['text'])
            elif kind == 'insert_rows':
                ws.insert_rows(op['at'], op['count'])
            elif kind == 'delete_rows':
                ws.delete_rows(op['at'], op['count'])
            elif kind == 'insert_cols':
                ws.insert_cols(op['at'], op['count'])
            elif kind == 'delete_cols':
                ws.delete_cols(op['at'], op['count'])
            elif kind == 'format':
                _apply_format(ws, op)
        wb.save(tmp_name)
        _atomic_replace(tmp_name, path)
    except BaseException:
        try:
            os.unlink(tmp_name)
        except OSError:
            pass
        raise
    finally:
        try:
            wb.close()
        except Exception:
            pass
    return made


def _encode_row(fields: Sequence[str], delimiter: str) -> str:
    buf = io.StringIO(newline='')
    csv.writer(buf, delimiter=delimiter, lineterminator='').writerow(list(fields))
    return buf.getvalue()


def apply_ops_delimited(
    path: Path,
    ops: Any,
    delimiter: str,
    *,
    backup: bool = True,
    backup_dir: Optional[Path] = None,
) -> Optional[Path]:
    """Replay the same edit journal against a csv/tsv file.

    Only lines the journal actually touches are re-serialised; every other line
    is written back byte for byte. Re-encoding the whole file would silently
    normalise quoting and spacing on rows the user never looked at.
    """
    path = Path(path)
    clean = validate_ops(ops)
    if not clean:
        return None
    made = (make_backup(path, folder=backup_dir, versioned=bool(backup_dir))
            if backup else None)

    # Decoded from bytes rather than read as text: that keeps the original
    # line endings visible (text mode would translate them and we would
    # silently rewrite a CRLF file as LF), and it is what tells us which
    # encoding to write back in.
    raw_text, encoding = decode_delimited(path)
    newline = '\r\n' if '\r\n' in raw_text else '\n'
    trailing_newline = raw_text.endswith(('\n', '\r'))
    raw_lines = raw_text.splitlines()

    # One entry per line: the original text, its parsed fields, and whether the
    # journal changed it. Untouched entries keep `raw`.
    rows: List[Dict[str, Any]] = []
    for line in raw_lines:
        fields = next(csv.reader([line], delimiter=delimiter), [])
        rows.append({'raw': line, 'fields': list(fields), 'dirty': False})

    def _ensure(r_count: int, c_count: int) -> None:
        while len(rows) < r_count:
            rows.append({'raw': '', 'fields': [], 'dirty': True})
        row = rows[r_count - 1] if r_count else None
        if row is not None and len(row['fields']) < c_count:
            row['fields'].extend([''] * (c_count - len(row['fields'])))
            row['dirty'] = True

    if any(op['op'] == 'format' for op in clean):
        raise SpreadsheetError(
            'A csv file holds text, not formatting. Save it as .xlsx to keep '
            'colours and number formats.')

    for op in clean:
        kind = op['op']
        if kind == 'set':
            _ensure(op['row'], op['col'])
            row = rows[op['row'] - 1]
            if row['fields'][op['col'] - 1] != op['text']:
                row['fields'][op['col'] - 1] = op['text']
                row['dirty'] = True
        elif kind == 'insert_rows':
            _ensure(max(op['at'] - 1, 1), 1)
            width = max([1] + [len(r['fields']) for r in rows])
            for _ in range(op['count']):
                rows.insert(op['at'] - 1,
                            {'raw': '', 'fields': [''] * width, 'dirty': True})
        elif kind == 'delete_rows':
            start = op['at'] - 1
            del rows[start:start + op['count']]
        elif kind == 'insert_cols':
            for row in rows:
                for _ in range(op['count']):
                    row['fields'].insert(min(op['at'] - 1, len(row['fields'])), '')
                row['dirty'] = True
        elif kind == 'delete_cols':
            start = op['at'] - 1
            for row in rows:
                if start < len(row['fields']):
                    del row['fields'][start:start + op['count']]
                    row['dirty'] = True

    out_lines = [
        _encode_row(r['fields'], delimiter) if r['dirty'] else r['raw']
        for r in rows
    ]
    text = newline.join(out_lines)
    if trailing_newline and text:
        text += newline

    payload = encode_delimited(text, encoding)
    tmp_fd, tmp_name = tempfile.mkstemp(dir=str(path.parent), prefix='.dsheet-', suffix=path.suffix)
    try:
        with os.fdopen(tmp_fd, 'wb') as fh:
            fh.write(payload)
        _atomic_replace(tmp_name, path)
    except BaseException:
        try:
            os.unlink(tmp_name)
        except OSError:
            pass
        raise
    return made


# ---------------------------------------------------------------------------
# CSS
# ---------------------------------------------------------------------------

# The tab enforces `.calc-tab * { overflow-x:hidden !important }` and
# `.calc-right * { max-width:100% !important }`. Both have to be beaten by
# specificity, not source order, because they carry !important.
#
# The overflow override is not cosmetic: an element with overflow-x:hidden and
# overflow-y:visible has its overflow-y computed to auto, which turns <table>,
# <thead>, <tbody> and <tr> into scroll containers -- and position:sticky
# resolves against the nearest scroll container, so the frozen header would
# stick to its own row instead of the viewport.
GRID_CSS = (
    '.calc-tab .dsheet-root { display:flex !important; flex-direction:column !important;'
    ' height:100% !important; min-height:0 !important; max-width:none !important;'
    ' overflow:hidden !important; }'
    '.calc-tab .dsheet-scroll { flex:1 1 0 !important; min-height:0 !important;'
    ' overflow:auto !important; max-width:none !important; position:relative !important;'
    ' border:1px solid #ddd; background:#fff; outline:none; }'
    '.calc-tab .dsheet-scroll table, .calc-tab .dsheet-scroll thead,'
    ' .calc-tab .dsheet-scroll tbody, .calc-tab .dsheet-scroll tr'
    ' { overflow:visible !important; max-width:none !important; }'
    '.calc-tab .dsheet-scroll col, .calc-tab .dsheet-scroll th,'
    ' .calc-tab .dsheet-scroll td { max-width:none !important; }'
    '.calc-tab .dsheet-scroll th, .calc-tab .dsheet-scroll td'
    ' { overflow:hidden !important; white-space:nowrap !important; text-overflow:ellipsis; }'
    '.calc-tab .dsheet-scroll td[contenteditable="true"]'
    ' { overflow:visible !important; white-space:pre !important; text-overflow:clip; }'
    '.calc-tab .dsheet-bar, .calc-tab .dsheet-tabs, .calc-tab .dsheet-foot'
    ' { max-width:none !important; overflow-x:auto !important; }'

    '.dsheet-root { font-family:-apple-system,Segoe UI,Roboto,Arial,sans-serif; }'
    '.dsheet-bar { display:flex; flex-wrap:wrap; align-items:center; gap:4px;'
    ' padding:4px 6px; border:1px solid #ddd; border-bottom:none; background:#f4f6f8;'
    ' flex:0 0 auto; }'
    '.dsheet-btn { font-size:11px; padding:2px 8px; height:22px; line-height:1;'
    ' border:1px solid #c3c9d0; border-radius:3px; background:#fff; cursor:pointer; color:#333; }'
    # One box, one baseline, whatever the letter in it is doing.
    '.dsheet-mark { display:inline-flex; align-items:center;'
    ' justify-content:center; width:24px; padding:0; font-size:13px; }'
    '.dsheet-b { font-weight:700; }'
    '.dsheet-i { font-style:italic; font-family:Georgia,serif; }'
    '.dsheet-u { text-decoration:underline; text-underline-offset:2px; }'
    '.dsheet-btn:hover:not(:disabled) { background:#e8f0fe; border-color:#7aa7e8; }'
    '.dsheet-btn:disabled { opacity:0.45; cursor:default; }'
    '.dsheet-btn.dsheet-primary { background:#1976d2; border-color:#1565c0; color:#fff; }'
    '.dsheet-btn.dsheet-primary:disabled { background:#b8c4d0; border-color:#b0bcc8; color:#fff; }'
    '.dsheet-sep { display:inline-block; width:1px; height:16px; background:#ccd2d8; margin:0 4px; }'
    '.dsheet-filter { font-size:11px; height:22px; padding:0 6px; width:140px;'
    ' border:1px solid #c3c9d0; border-radius:3px; }'
    '.dsheet-addr { font-family:monospace; font-size:11px; color:#555; min-width:42px; }'
    '.dsheet-status { font-size:11px; color:#666; }'
    '.dsheet-status.dsheet-warn { color:#b26a00; font-weight:600; }'
    '.dsheet-lossy { font-size:11px; color:#8a4b00; background:#fff4e0;'
    ' border:1px solid #f0d3a8; border-bottom:none; padding:3px 8px; flex:0 0 auto; }'
    '.dsheet-tabs { display:flex; gap:3px; padding:3px 6px; border-left:1px solid #ddd;'
    ' border-right:1px solid #ddd; background:#eef1f4; flex:0 0 auto; }'
    '.dsheet-tab { font-size:11px; padding:2px 9px; border:1px solid #c8ced4;'
    ' border-bottom:none; border-radius:3px 3px 0 0; background:#fafbfc; cursor:pointer;'
    ' white-space:nowrap; }'
    '.dsheet-tab.dsheet-tab-on { background:#fff; font-weight:600; border-color:#7aa7e8; }'
    '.dsheet-tab-add, .dsheet-tab-rename, .dsheet-tab-drop { color:#2b5c9b; }'
    '.dsheet-tab-drop { color:#b3261e; }'
    '.dsheet-name-input { font-size:11px; padding:2px 6px; border:1px solid #7aa7e8;'
    ' border-radius:3px; min-width:120px; }'

    '.dsheet { border-collapse:separate; border-spacing:0; table-layout:fixed; font-size:12px; }'
    '.dsheet th, .dsheet td { padding:2px 5px; height:21px; vertical-align:middle;'
    ' box-shadow:inset -1px 0 0 #e3e6e9, inset 0 -1px 0 #e3e6e9; }'
    '.dsheet thead th { position:sticky; top:0; z-index:3; background:#eef1f4;'
    ' font-weight:600; text-align:center; color:#444; user-select:none; cursor:default; }'
    '.dsheet tbody th { position:sticky; left:0; z-index:2; background:#eef1f4;'
    ' font-weight:400; text-align:center; color:#666; user-select:none; cursor:default; }'
    '.dsheet thead th.dsheet-corner { left:0; z-index:4; }'
    '.dsheet td { background:#fff; cursor:cell; }'
    '.dsheet td.dsheet-num { text-align:right; }'
    '.dsheet td.dsheet-sel { background:#e3f0fd; }'
    '.dsheet td.dsheet-cur { outline:2px solid #1976d2; outline-offset:-2px; position:relative; z-index:1; }'
    '.dsheet td.dsheet-hit { outline:2px solid #ff9800; outline-offset:-2px; }'
    '.dsheet-fills { display:inline-flex; gap:4px; align-items:center;'
    ' margin:0 6px; }'
    '.dsheet-swatch { width:22px; height:22px; border:1px solid #b7bec6;'
    ' border-radius:3px; cursor:pointer; display:inline-block; }'
    '.dsheet-swatch:hover { outline:2px solid #1565c0; outline-offset:1px; }'
    '.dsheet-swatch-none { background:#fff; color:#b3261e; font-size:15px;'
    ' line-height:20px; text-align:center; }'
    '.dsheet-fill { position:absolute; width:7px; height:7px; background:#1565c0;'
    ' border:1px solid #fff; cursor:crosshair; z-index:4; display:none; }'
    '.dsheet td.dsheet-fill-target { background:#e3f0ff;'
    ' outline:1px dashed #1565c0; outline-offset:-1px; }'
    '.dsheet td.dsheet-dirty { background:#fff6cc; }'
    '.dsheet td.dsheet-dirty.dsheet-sel { background:#f3e9b4; }'
    '.dsheet td[contenteditable="true"] { background:#fff; outline:2px solid #1976d2;'
    ' outline-offset:-2px; cursor:text; z-index:6; position:relative; }'
    '.dsheet-grip { position:absolute; top:0; right:0; width:6px; height:100%;'
    ' cursor:col-resize; z-index:5; }'
    '.dsheet-grip-row { position:absolute; left:0; bottom:0; width:100%; height:5px;'
    ' cursor:row-resize; z-index:5; }'
    '.dsheet thead th, .dsheet tbody th { position:sticky; }'
    '.dsheet-menu { position:fixed; z-index:9999; background:#fff; border:1px solid #c3c9d0;'
    ' border-radius:4px; box-shadow:0 3px 10px rgba(0,0,0,0.18); padding:3px 0;'
    ' font-size:12px; min-width:190px; }'
    '.dsheet-menu div { padding:4px 14px; cursor:pointer; white-space:nowrap; }'
    '.dsheet-menu div:hover { background:#e8f0fe; }'
    '.dsheet-foot { flex:0 0 auto; display:flex; align-items:center; gap:8px;'
    ' padding:3px 6px; border:1px solid #ddd; border-top:none; background:#f4f6f8;'
    ' font-size:11px; color:#666; }'
    '.dsheet-foot .dsheet-warn { color:#b26a00; }'
)


# ---------------------------------------------------------------------------
# HTML
# ---------------------------------------------------------------------------

def _attr(value: Any) -> str:
    return _html.escape(str(value if value is not None else ''), quote=True)


def _looks_numeric(text: str) -> bool:
    if not text:
        return False
    return bool(_INT_RE.match(text) or _FLOAT_RE.match(text) or _SCI_INT_RE.match(text))


def render_grid_html(
    sheet: SheetData,
    *,
    sheet_names: Sequence[str] = (),
    token: str = '',
    kind: str = 'xlsx',
    path: str = '',
    delimiter: str = '',
    editable: bool = True,
    dirty: Optional[Set[Tuple[int, int]]] = None,
    pending: int = 0,
    col_px: Optional[Sequence[int]] = None,
    lossy_note: str = '',
    scroll_top: int = 0,
    cursor: Optional[Tuple[int, int]] = None,
    office: bool = False,
) -> str:
    """Build the complete grid markup for the tab's content area.

    Cells carry no inline styles and no per-cell data attributes: the grid HTML
    is re-sent over the widget comm on every render, so a plain ``<td>`` keeps a
    few thousand cells to a couple of hundred KB. The browser derives addresses
    from ``cellIndex`` and the row's ``data-r``.
    """
    dirty = dirty or set()
    widths = list(col_px) if col_px and len(col_px) == sheet.n_cols else list(sheet.col_widths)
    if len(widths) < sheet.n_cols:
        widths.extend([DEFAULT_COL_WIDTH] * (sheet.n_cols - len(widths)))
    total_px = ROW_HEADER_WIDTH + sum(int(w) for w in widths[:sheet.n_cols])

    # Where to put the user back when the grid has to be rebuilt. Absent on a
    # first render, where the top-left cell is the right answer.
    cursor_attr = (
        f' data-cursor="{int(cursor[0])},{int(cursor[1])}"' if cursor else ''
    )
    out: List[str] = [f'<style>{GRID_CSS}</style>']
    out.append(
        '<div class="dsheet-root"'
        f' data-token="{_attr(token)}"'
        f' data-kind="{_attr(kind)}"'
        f' data-path="{_attr(path)}"'
        f' data-sheet="{_attr(sheet.name)}"'
        f' data-delim="{_attr(delimiter)}"'
        f' data-rowoffset="{sheet.row_offset}"'
        f' data-pending="{int(pending)}"'
        f' data-scrolltop="{int(scroll_top)}"'
        f'{cursor_attr}'
        f' data-office="{"1" if office else "0"}"'
        f' data-editable="{"1" if editable else "0"}">'
    )

    # --- toolbar ---
    out.append('<div class="dsheet-bar">')
    if editable:
        dis = '' if pending else ' disabled'
        out.append(f'<button class="dsheet-btn dsheet-primary dsheet-save"{dis}>Save</button>')
        out.append(f'<button class="dsheet-btn dsheet-discard"{dis}>Discard</button>')
        out.append('<span class="dsheet-sep"></span>')
        if kind == 'xlsx':
            # A csv holds text, not formatting, so these are not offered there.
            # The letters are plain and the button carries the styling. A
            # <b>, an <i> and a <u> inside three buttons have three
            # different line boxes, so the three buttons stopped lining up.
            out.append('<button class="dsheet-btn dsheet-mark dsheet-b"'
                       ' title="Bold (Ctrl+B)">B</button>')
            out.append('<button class="dsheet-btn dsheet-mark dsheet-i"'
                       ' title="Italic (Ctrl+I)">I</button>')
            out.append('<button class="dsheet-btn dsheet-mark dsheet-u"'
                       ' title="Underline (Ctrl+U)">U</button>')
            out.append('<span class="dsheet-fills">')
            for label, colour in FILL_COLOURS:
                if colour:
                    out.append(
                        f'<span class="dsheet-swatch" data-fill="{colour}"'
                        f' title="{_attr(label)}"'
                        f' style="background:#{colour}"></span>')
                else:
                    out.append(
                        '<span class="dsheet-swatch dsheet-swatch-none"'
                        f' data-fill="" title="{_attr(label)}">&times;</span>')
            out.append('</span>')
            # No control for the number format yet. Setting one changes how a
            # number should read, and the browser cannot re-read it: picking
            # a format did nothing visible until the file had been saved and
            # opened again, which is worse than not offering it. A format a
            # workbook already carries is still read and shown.
            out.append('<span class="dsheet-sep"></span>')
        if office:
            out.append('<button class="dsheet-btn dsheet-undo" disabled'
                       ' title="Undo (Ctrl+Z)">&#8630;</button>')
            out.append('<button class="dsheet-btn dsheet-redo" disabled'
                       ' title="Redo (Ctrl+Shift+Z)">&#8631;</button>')
            out.append('<span class="dsheet-sep"></span>')
        out.append('<button class="dsheet-btn" data-act="insert_rows">+ Row</button>')
        out.append('<button class="dsheet-btn" data-act="delete_rows">&minus; Row</button>')
        out.append('<button class="dsheet-btn" data-act="insert_cols">+ Column</button>')
        out.append('<button class="dsheet-btn" data-act="delete_cols">&minus; Column</button>')
        out.append('<span class="dsheet-sep"></span>')
    out.append('<span class="dsheet-addr">A1</span>')
    out.append('<input class="dsheet-filter" placeholder="Filter…" spellcheck="false">')
    out.append('<span class="dsheet-status"></span>')
    out.append('</div>')

    if lossy_note:
        out.append(f'<div class="dsheet-lossy">{_html.escape(lossy_note)}</div>')

    # --- sheet tabs ---
    # Shown for a single sheet too, in a workbook: that is where the control
    # for adding the second one lives, and a strip that appears only once
    # there is already more than one is a strip nobody finds.
    names = [n for n in sheet_names if n]
    if names and kind == 'xlsx':
        out.append('<div class="dsheet-tabs">')
        for name in names:
            on = ' dsheet-tab-on' if name == sheet.name else ''
            out.append(
                f'<span class="dsheet-tab{on}" data-sheet="{_attr(name)}"'
                f' title="{_attr(name)}">{_html.escape(name)}</span>'
            )
        if editable:
            out.append('<span class="dsheet-tab dsheet-tab-add"'
                       ' title="New sheet">+</span>')
            out.append('<span class="dsheet-tab dsheet-tab-rename"'
                       ' title="Rename this sheet">Rename</span>')
            if len(names) > 1:
                # The last sheet cannot go: a workbook has to have one.
                out.append('<span class="dsheet-tab dsheet-tab-drop"'
                           ' title="Delete this sheet">&minus;</span>')
        out.append('</div>')

    # --- grid ---
    out.append('<div class="dsheet-scroll" tabindex="0">')
    out.append(f'<table class="dsheet" spellcheck="false" style="width:{total_px}px;">')
    out.append(f'<colgroup><col style="width:{ROW_HEADER_WIDTH}px;">')
    for w in widths[:sheet.n_cols]:
        out.append(f'<col style="width:{int(w)}px;">')
    out.append('</colgroup>')

    out.append('<thead><tr><th class="dsheet-corner"></th>')
    for c_idx in range(sheet.n_cols):
        out.append(f'<th>{col_letter(c_idx)}<span class="dsheet-grip"></span></th>')
    out.append('</tr></thead><tbody>')

    formulas = sheet.formulas
    for r_idx, row in enumerate(sheet.values):
        abs_row = sheet.row_offset + r_idx + 1
        out.append(f'<tr data-r="{abs_row}"><th>{abs_row}<span class="dsheet-grip-row"></span></th>')
        f_row = formulas[r_idx] if r_idx < len(formulas) else None
        for c_idx, cell in enumerate(row):
            formula = f_row[c_idx] if f_row and c_idx < len(f_row) else ''
            classes = []
            if not formula and _looks_numeric(cell):
                classes.append('dsheet-num')
            if (r_idx, c_idx) in dirty:
                classes.append('dsheet-dirty')
            cls = f' class="{" ".join(classes)}"' if classes else ''
            f_attr = f' data-f="{_attr(formula)}"' if formula else ''
            style = {}
            if r_idx < len(sheet.styles) and c_idx < len(sheet.styles[r_idx]):
                style = sheet.styles[r_idx][c_idx] or {}
            css = style_css(style)
            s_attr = f' style="{css}"' if css else ''
            bg = style.get('bg')
            bg_attr = f' data-bg="{_attr(bg)}"' if bg else ''
            fmt = style.get('fmt')
            n_attr = f' data-n="{_attr(fmt)}"' if fmt else ''
            out.append(f'<td{cls}{s_attr}{bg_attr}{f_attr}{n_attr}>{_html.escape(cell)}</td>')
        out.append('</tr>')
    out.append('</tbody></table></div>')

    # --- footer ---
    first = sheet.row_offset + 1
    last = sheet.row_offset + sheet.n_rows
    out.append('<div class="dsheet-foot">')
    out.append(
        f'<span>Rows {first}–{last} of {max(sheet.total_rows, last)}'
        f' &middot; {sheet.total_cols} columns</span>'
    )
    if sheet.row_offset > 0:
        out.append('<button class="dsheet-btn" data-page="prev">&laquo; previous</button>')
    if sheet.total_rows > last:
        out.append('<button class="dsheet-btn" data-page="next">next &raquo;</button>')
    if sheet.truncated_cols:
        out.append(
            f'<span class="dsheet-warn">Only the first {sheet.n_cols} columns are shown.</span>'
        )
    if sheet.has_formulas:
        out.append(
            '<span class="dsheet-warn">Formulas are recalculated when Excel opens the file.</span>'
        )
    out.append('</div></div>')
    return ''.join(out)


# ---------------------------------------------------------------------------
# Browser side
# ---------------------------------------------------------------------------

_GRID_JS_TEMPLATE = r"""
(function(){
  var SCOPE = __SCOPE__;
  var TOKEN = __TOKEN__;
  var tries = 0;

  function boot(){
    var root = document.querySelector('.' + SCOPE);
    var wrap = root ? root.querySelector('.dsheet-root[data-token="' + TOKEN + '"]') : null;
    if (!wrap) {
      if (++tries < 40) { setTimeout(boot, 50); }
      return;
    }
    if (wrap.dataset.bound === '1') return;
    wrap.dataset.bound = '1';
    install(root, wrap);
  }

  function install(root, wrap){
  var table = wrap.querySelector('table.dsheet');
  var scroll = wrap.querySelector('.dsheet-scroll');
  var tbody = table ? table.querySelector('tbody') : null;
  var thead = table ? table.querySelector('thead tr') : null;
  var colgroup = table ? table.querySelector('colgroup') : null;
  if (!table || !scroll || !tbody || !thead) return;

  var editable = wrap.dataset.editable === '1';
  /* Office keeps the viewport still when a whole column or row is picked;
     the calculations browser keeps the behaviour it always had. */
  var OFFICE = wrap.dataset.office === '1';
  var pending = parseInt(wrap.dataset.pending || '0', 10) || 0;
  var rowOffset = parseInt(wrap.dataset.rowoffset || '0', 10) || 0;
  var sortState = null;
  var filterActive = false;
  var editing = null;
  var cur = {r: 0, c: 1};
  var anchor = {r: 0, c: 1};
  var menu = null;
  var copyHandled = false;

  var saveBtn = wrap.querySelector('.dsheet-save');
  var discardBtn = wrap.querySelector('.dsheet-discard');
  var undoBtn = wrap.querySelector('.dsheet-undo');
  var redoBtn = wrap.querySelector('.dsheet-redo');
  var statusEl = wrap.querySelector('.dsheet-status');
  var addrEl = wrap.querySelector('.dsheet-addr');
  var filterEl = wrap.querySelector('.dsheet-filter');

  /* ---------- widget bridge ---------- */
  function setField(cls, value){
    var el = root.querySelector('.' + cls + ' textarea, .' + cls + ' input');
    if (!el) return false;
    var str = String(value == null ? '' : value);
    var desc = Object.getOwnPropertyDescriptor(Object.getPrototypeOf(el), 'value')
      || Object.getOwnPropertyDescriptor(HTMLTextAreaElement.prototype, 'value')
      || Object.getOwnPropertyDescriptor(HTMLInputElement.prototype, 'value');
    if (desc && desc.set) { desc.set.call(el, str); } else { el.value = str; }
    el.dispatchEvent(new Event('input', {bubbles: true}));
    el.dispatchEvent(new Event('change', {bubbles: true}));
    return true;
  }
  function clickTrigger(cls){
    var node = root.querySelector('.' + cls);
    if (!node) return false;
    var btn = node.tagName === 'BUTTON' ? node : node.querySelector('button');
    if (!btn) return false;
    btn.click();
    return true;
  }
  function send(action, ops, extra){
    var payload = {
      action: action, token: TOKEN,
      sheet: wrap.dataset.sheet, kind: wrap.dataset.kind,
      ops: ops || [], cols: colWidths(), scroll: scroll.scrollTop,
      /* Where the user was. Anything that has to rebuild the grid puts them
         back here, instead of at A1 with the selection gone. */
      cur: [cur.r, cur.c]
    };
    if (extra) { for (var k in extra) { payload[k] = extra[k]; } }
    if (!setField('calc-sheet-payload', JSON.stringify(payload))) return;
    clickTrigger('calc-sheet-action');
  }
  /* Every committed change goes to Python immediately, so clicking another
     file in the browser (which re-renders this widget) cannot lose it. */
  function push(ops){
    pending += ops.length;
    reflectPending();
    send('edit', ops);
  }

  /* ---------- undo / redo ----------
     An undo is applied as a further change rather than by rewinding the
     journal: the file may already have been saved, and in a spreadsheet
     undoing after a save is allowed and marks the file changed again.
     Each step therefore knows how to do itself and how to take itself
     back, and both go to Python the ordinary way. */
  var history = [], future = [], replaying = false;
  var HISTORY_MAX = 300;

  function remember(step){
    if (replaying) return;
    history.push(step);
    if (history.length > HISTORY_MAX) history.shift();
    future.length = 0;          /* a new change ends the redo branch */
    reflectHistory();
  }
  function quietly(fn){
    replaying = true;
    try { fn(); } finally { replaying = false; }
  }
  function undo(){
    var step = history.pop();
    if (!step) return;
    quietly(step.undo);
    future.push(step);
    reflectHistory();
  }
  function redo(){
    var step = future.pop();
    if (!step) return;
    quietly(step.redo);
    history.push(step);
    reflectHistory();
  }
  function reflectHistory(){
    if (undoBtn) undoBtn.disabled = history.length === 0;
    if (redoBtn) redoBtn.disabled = future.length === 0;
  }
  function cellText(td){
    return td.getAttribute('data-f') || td.textContent;
  }
  /* Write cells and keep what was there, so the step can take itself back.
     An entry may carry its own `was`: a cell being edited is contenteditable,
     so by the time the edit is committed the DOM already holds the new text
     and reading it back would record the change as its own undo. */
  function writeCells(entries){
    if (!entries.length) return;
    var before = entries.map(function(e){
      return {td: e.td, text: e.was === undefined ? cellText(e.td) : e.was};
    });
    function apply(list){
      push(list.map(function(e){ return applyText(e.td, e.text); }));
    }
    apply(entries);
    remember({
      undo: function(){ apply(before); },
      redo: function(){ apply(entries); }
    });
  }

  /* ---------- geometry ---------- */
  function colWidths(){
    var out = [], cols = colgroup ? colgroup.querySelectorAll('col') : [];
    for (var i = 1; i < cols.length; i++) out.push(parseInt(cols[i].style.width, 10) || 110);
    return out;
  }
  function retotal(){
    var cols = colgroup.querySelectorAll('col'), total = 0;
    for (var i = 0; i < cols.length; i++) total += parseInt(cols[i].style.width, 10) || 0;
    table.style.width = total + 'px';
  }
  function cellAt(r, c){
    var tr = tbody.rows[r];
    if (!tr || c < 1 || c >= tr.cells.length) return null;
    return tr.cells[c];
  }
  function colCount(){ return thead.cells.length - 1; }
  function rowIndexOf(tr){ return Array.prototype.indexOf.call(tbody.rows, tr); }
  function colName(c){
    var n = c, s = '';
    while (n > 0) { var rem = (n - 1) % 26; s = String.fromCharCode(65 + rem) + s; n = Math.floor((n - 1) / 26); }
    return s;
  }

  /* ---------- formatting ---------- */
  /* Applied to the selection and shown at once, then sent. Waiting for the
     kernel to answer before the cells change would make every click feel
     like a request rather than like formatting. */
  function readFormat(td){
    return {
      bold: td.style.fontWeight === '700',
      italic: td.style.fontStyle === 'italic',
      underline: td.style.textDecoration === 'underline',
      fill: td.getAttribute('data-bg') || '',
      number_format: td.getAttribute('data-n') || 'General'
    };
  }

  function paintCell(td, change){
    if ('bold' in change) td.style.fontWeight = change.bold ? '700' : '';
    if ('italic' in change) td.style.fontStyle = change.italic ? 'italic' : '';
    if ('underline' in change) {
      td.style.textDecoration = change.underline ? 'underline' : '';
    }
    if ('fill' in change) {
      if (change.fill) {
        td.style.background = '#' + change.fill;
        td.setAttribute('data-bg', change.fill);
      } else {
        td.style.background = '';
        td.removeAttribute('data-bg');
      }
    }
    if ('number_format' in change) {
      if (change.number_format && change.number_format !== 'General') {
        td.setAttribute('data-n', change.number_format);
      } else {
        td.removeAttribute('data-n');
      }
    }
  }

  /* One op per cell, carrying only the keys that changed. A range is rarely
     uniform -- half of it may already be bold -- so putting it back needs
     each cell's own previous state, not the range's. */
  function formatOps(cells, pick){
    var ops = [];
    for (var i = 0; i < cells.length; i++) {
      var entry = cells[i];
      var op = {op: 'format', r1: entry.row, c1: entry.col,
                r2: entry.row, c2: entry.col};
      var values = pick(entry);
      var touched = false;
      for (var key in values) { op[key] = values[key]; touched = true; }
      if (touched) ops.push(op);
    }
    return ops;
  }

  function applyFormat(change){
    if (!editable) return;
    var s = selRange();
    var cells = [];
    for (var r = s.r1; r <= s.r2; r++) {
      var tr = tbody.rows[r];
      if (!tr) continue;
      for (var c = s.c1; c <= s.c2; c++) {
        var td = cellAt(r, c);
        if (!td) continue;
        var was = readFormat(td);
        var before = {};
        for (var key in change) before[key] = was[key];
        cells.push({td: td, row: parseInt(tr.dataset.r, 10), col: c,
                    before: before});
      }
    }
    if (!cells.length) return;

    function run(pick){
      for (var i = 0; i < cells.length; i++) paintCell(cells[i].td, pick(cells[i]));
      var ops = formatOps(cells, pick);
      if (ops.length) push(ops);
    }

    run(function(){ return change; });
    remember({
      undo: function(){ run(function(entry){ return entry.before; }); },
      redo: function(){ run(function(){ return change; }); }
    });
  }

  function selectionHas(prop){
    var s = selRange();
    var td = cellAt(s.r1, s.c1);
    return td ? !!readFormat(td)[prop] : false;
  }

  var boldBtn = wrap.querySelector('.dsheet-b');
  var italicBtn = wrap.querySelector('.dsheet-i');
  var underlineBtn = wrap.querySelector('.dsheet-u');
  if (boldBtn) boldBtn.addEventListener('click', function(){
    applyFormat({bold: !selectionHas('bold')});
  });
  if (italicBtn) italicBtn.addEventListener('click', function(){
    applyFormat({italic: !selectionHas('italic')});
  });
  if (underlineBtn) underlineBtn.addEventListener('click', function(){
    applyFormat({underline: !selectionHas('underline')});
  });
  Array.prototype.forEach.call(wrap.querySelectorAll('.dsheet-swatch'),
    function(swatch){
      swatch.addEventListener('click', function(){
        applyFormat({fill: swatch.getAttribute('data-fill') || ''});
      });
    });

  /* ---------- fill handle ---------- */
  /* The square at the bottom right of the selection. Dragging it is how a
     spreadsheet copies a cell down a column, and it is the difference
     between entering a formula once and entering it three hundred times. */
  var fillHandle = null;
  if (editable) {
    fillHandle = document.createElement('div');
    fillHandle.className = 'dsheet-fill';
    fillHandle.title = 'Drag to fill down or across';
    scroll.appendChild(fillHandle);
  }
  var filling = null;

  function placeHandle(){
    if (!fillHandle) return;
    var s = selRange();
    var td = cellAt(s.r2, s.c2);
    if (!td || filling) { if (!filling) fillHandle.style.display = 'none'; return; }
    fillHandle.style.display = 'block';
    fillHandle.style.left = (td.offsetLeft + td.offsetWidth - 5) + 'px';
    fillHandle.style.top = (td.offsetTop + td.offsetHeight - 5) + 'px';
  }

  function clearFillTargets(){
    Array.prototype.forEach.call(
      tbody.querySelectorAll('td.dsheet-fill-target'),
      function(td){ td.classList.remove('dsheet-fill-target'); });
  }

  function markFillTargets(){
    clearFillTargets();
    if (!filling || (!filling.rows && !filling.cols)) return;
    var s = filling.from;
    for (var r = s.r1; r <= s.r2 + filling.rows; r++) {
      for (var c = s.c1; c <= s.c2 + filling.cols; c++) {
        if (r <= s.r2 && c <= s.c2) continue;
        var td = cellAt(r, c);
        if (td) td.classList.add('dsheet-fill-target');
      }
    }
  }

  if (fillHandle) {
    fillHandle.addEventListener('mousedown', function(e){
      e.preventDefault();
      e.stopPropagation();
      if (editing) commitEdit();
      filling = {from: selRange(), rows: 0, cols: 0};
      fillHandle.style.display = 'none';
    });
  }

  scroll.addEventListener('mousemove', function(e){
    if (!filling) return;
    var td = e.target.closest ? e.target.closest('td') : null;
    if (!td || td.parentNode.parentNode !== tbody) return;
    var r = rowIndexOf(td.parentNode), c = td.cellIndex;
    var s = filling.from;
    /* One axis at a time, like the handle in a spreadsheet: whichever the
       pointer has travelled further along. */
    var down = Math.max(0, r - s.r2), across = Math.max(0, c - s.c2);
    if (down >= across) { filling.rows = down; filling.cols = 0; }
    else { filling.rows = 0; filling.cols = across; }
    markFillTargets();
  }, false);

  document.addEventListener('mouseup', function(){
    if (!filling) return;
    var job = filling;
    filling = null;
    clearFillTargets();
    placeHandle();
    if (!job.rows && !job.cols) return;
    var s = job.from;
    var block = [];
    for (var r = s.r1; r <= s.r2; r++) {
      var row = [];
      for (var c = s.c1; c <= s.c2; c++) {
        var td = cellAt(r, c);
        row.push(td ? cellText(td) : '');
      }
      block.push(row);
    }
    var firstRow = tbody.rows[s.r1];
    send('fill', [], {
      block: block,
      rows: job.rows,
      cols: job.cols,
      at: [parseInt(firstRow.dataset.r, 10) || (s.r1 + 1 + rowOffset), s.c1]
    });
    /* Extend the selection over what was just filled, as a spreadsheet does. */
    anchor = {r: s.r1, c: s.c1};
    moveTo(s.r2 + job.rows, s.c2 + job.cols, true, true);
  }, false);

  /* Worked-out results. Not an edit: the cell keeps its formula, is not
     marked changed, and undo has nothing to take back -- the user did not
     type these. */
  wrap.__dsheetShow = function(cells){
    for (var i = 0; i < cells.length; i++) {
      var absRow = cells[i][0], col = cells[i][1], text = cells[i][2];
      for (var j = 0; j < tbody.rows.length; j++) {
        if (parseInt(tbody.rows[j].dataset.r, 10) !== absRow) continue;
        var td = cellAt(j, col);
        if (td && td.getAttribute('data-f')) td.textContent = text;
        break;
      }
    }
  };

  /* Values worked out by the kernel, written in as one step so one undo
     takes the whole fill back. */
  wrap.__dsheetApply = function(cells){
    var writes = [];
    for (var i = 0; i < cells.length; i++) {
      var absRow = cells[i][0], col = cells[i][1], text = cells[i][2];
      var target = -1;
      for (var j = 0; j < tbody.rows.length; j++) {
        if (parseInt(tbody.rows[j].dataset.r, 10) === absRow) { target = j; break; }
      }
      if (target < 0) continue;
      var td = cellAt(target, col);
      if (td) writes.push({td: td, text: text});
    }
    writeCells(writes);
  };

  /* ---------- selection ---------- */
  function paint(){
    var r1 = Math.min(anchor.r, cur.r), r2 = Math.max(anchor.r, cur.r);
    var c1 = Math.min(anchor.c, cur.c), c2 = Math.max(anchor.c, cur.c);
    for (var i = 0; i < tbody.rows.length; i++) {
      var tr = tbody.rows[i];
      for (var j = 1; j < tr.cells.length; j++) {
        var td = tr.cells[j];
        td.classList.toggle('dsheet-sel', i >= r1 && i <= r2 && j >= c1 && j <= c2);
        td.classList.toggle('dsheet-cur', i === cur.r && j === cur.c);
      }
    }
    var tr0 = tbody.rows[cur.r];
    if (tr0 && addrEl) addrEl.textContent = colName(cur.c) + (tr0.dataset.r || '');
    placeHandle();
  }
  /* keepView: select without scrolling. Picking a whole column or row is a
     selection, not a journey -- the sheet has to stay where the eye left it. */
  function moveTo(r, c, extend, keepView){
    var maxR = tbody.rows.length - 1, maxC = colCount();
    r = Math.max(0, Math.min(maxR, r));
    c = Math.max(1, Math.min(maxC, c));
    if (tbody.rows[r] && tbody.rows[r].style.display === 'none') {
      var best = -1;
      for (var i = 0; i < tbody.rows.length; i++) {
        if (tbody.rows[i].style.display !== 'none') { if (i <= r || best < 0) best = i; }
      }
      if (best < 0) return;
      r = best;
    }
    cur = {r: r, c: c};
    if (!extend) anchor = {r: r, c: c};
    paint();
    var td = cellAt(r, c);
    if (td && !keepView) reveal(td);
  }
  function reveal(td){
    var headH = thead.getBoundingClientRect().height || 22;
    var headW = tbody.rows[0] ? tbody.rows[0].cells[0].offsetWidth : 54;
    var top = td.offsetTop, left = td.offsetLeft, h = td.offsetHeight, w = td.offsetWidth;
    if (top - headH < scroll.scrollTop) scroll.scrollTop = Math.max(0, top - headH);
    else if (top + h > scroll.scrollTop + scroll.clientHeight) scroll.scrollTop = top + h - scroll.clientHeight;
    if (left - headW < scroll.scrollLeft) scroll.scrollLeft = Math.max(0, left - headW);
    else if (left + w > scroll.scrollLeft + scroll.clientWidth) scroll.scrollLeft = left + w - scroll.clientWidth;
  }

  /* ---------- status ---------- */
  function reflectPending(){
    if (saveBtn) saveBtn.disabled = pending === 0;
    if (discardBtn) discardBtn.disabled = pending === 0;
    if (statusEl) {
      statusEl.textContent = pending
        ? (pending + (pending === 1 ? ' change' : ' changes') + ' not saved')
        : '';
      statusEl.classList.toggle('dsheet-warn', pending > 0);
    }
  }
  function flash(msg){
    if (!statusEl) return;
    statusEl.textContent = msg;
    statusEl.classList.add('dsheet-warn');
    setTimeout(reflectPending, 4000);
  }
  /* Saving is not a reason to rebuild the grid. The file is on disk and the
     cells on screen already show what was written, so all that changes here
     is the marks, the buttons and the message. Re-rendering would hand back
     a fresh table with the cursor on A1, the selection gone and the scroll
     wherever the new markup happened to measure -- which is not what
     pressing save in a spreadsheet does. */
  /* Go to a cell the kernel found. The row is the sheet's own number, so
     it is looked up rather than computed: the window on screen starts at
     an offset, and a filter can hide rows in between. Centred rather than
     merely revealed -- a hit at the edge of the view reads as "not found"
     until the eye finds it. */
  wrap.__dsheetGoto = function(rowNo, colNo){
    var target = -1;
    for (var i = 0; i < tbody.rows.length; i++) {
      if (parseInt(tbody.rows[i].dataset.r, 10) === rowNo) { target = i; break; }
    }
    if (target < 0) return false;
    var col = Math.max(1, Math.min(colCount(), colNo));
    anchor = {r: target, c: col};
    moveTo(target, col, false, true);
    var td = cellAt(target, col);
    if (td) {
      scroll.scrollTop = Math.max(
        0, td.offsetTop - Math.floor(scroll.clientHeight / 2));
      scroll.scrollLeft = Math.max(
        0, td.offsetLeft - Math.floor(scroll.clientWidth / 2));
      Array.prototype.forEach.call(
        tbody.querySelectorAll('td.dsheet-hit'),
        function(c){ c.classList.remove('dsheet-hit'); });
      td.classList.add('dsheet-hit');
    }
    return true;
  };
  wrap.__dsheetSaved = function(message){
    pending = 0;
    Array.prototype.forEach.call(
      tbody.querySelectorAll('td.dsheet-dirty'),
      function(td){ td.classList.remove('dsheet-dirty'); });
    reflectPending();
    if (statusEl && message) statusEl.textContent = message;
  };
  function structuralAllowed(){
    if (sortState || filterActive) {
      flash('Clear the sort or filter first - the target row is ambiguous otherwise.');
      return false;
    }
    return true;
  }

  /* ---------- in-cell editing ---------- */
  function beginEdit(td, seed){
    if (!editable || !td || editing) return;
    var formula = td.getAttribute('data-f') || '';
    var disp = td.textContent;
    editing = {td: td, disp: disp, formula: formula, source: formula || disp};
    td.setAttribute('contenteditable', 'true');
    td.textContent = (seed === null || seed === undefined) ? editing.source : seed;
    td.focus();
    var range = document.createRange();
    range.selectNodeContents(td);
    if (seed !== null && seed !== undefined) range.collapse(false);
    var s = window.getSelection();
    s.removeAllRanges();
    s.addRange(range);
  }
  /* Commit reports the change itself, so every exit path -- Enter, Tab, blur,
     or a mousedown on another cell -- lands in the journal exactly once. */
  function commitEdit(){
    if (!editing) return;
    var td = editing.td, source = editing.source, disp = editing.disp;
    var text = td.textContent.replace(/\u00a0/g, ' ');
    // Clear the guard first: dropping contenteditable blurs the cell
    // synchronously, and the focusout handler would otherwise re-enter here
    // and report the same change twice.
    editing = null;
    td.removeAttribute('contenteditable');
    if (text === source) {
      td.textContent = disp;
    } else {
      writeCells([{td: td, text: text, was: source}]);
    }
    scroll.focus({preventScroll: true});
  }
  function cancelEdit(){
    if (!editing) return;
    var td = editing.td, disp = editing.disp;
    editing = null;                      // see commitEdit: blur is synchronous
    td.removeAttribute('contenteditable');
    td.textContent = disp;
    scroll.focus({preventScroll: true});
  }
  function applyText(td, text){
    if (text.charAt(0) === '=') td.setAttribute('data-f', text);
    else td.removeAttribute('data-f');
    td.textContent = text;
    td.classList.add('dsheet-dirty');
    return {op: 'set', row: parseInt(td.parentNode.dataset.r, 10), col: td.cellIndex, text: text};
  }

  /* ---------- clipboard ---------- */
  function selRange(){
    return {
      r1: Math.min(anchor.r, cur.r), r2: Math.max(anchor.r, cur.r),
      c1: Math.min(anchor.c, cur.c), c2: Math.max(anchor.c, cur.c)
    };
  }
  function selToTsv(){
    var s = selRange(), out = [];
    for (var i = s.r1; i <= s.r2; i++) {
      if (tbody.rows[i].style.display === 'none') continue;
      var line = [];
      for (var j = s.c1; j <= s.c2; j++) {
        var td = cellAt(i, j);
        line.push(td ? (td.getAttribute('data-f') || td.textContent) : '');
      }
      out.push(line.join('\t'));
    }
    return out.join('\n');
  }
  function copyFallback(text){
    var ta = document.createElement('textarea');
    ta.value = text;
    ta.style.position = 'fixed';
    ta.style.opacity = '0';
    document.body.appendChild(ta);
    ta.select();
    try { document.execCommand('copy'); } catch (e) {}
    document.body.removeChild(ta);
  }
  function pasteTsv(text){
    if (!editable) return;
    var lines = String(text).replace(/\r\n/g, '\n').replace(/\r/g, '\n').split('\n');
    while (lines.length && lines[lines.length - 1] === '') lines.pop();
    if (!lines.length) return;
    var writes = [], wide = 1;
    for (var i = 0; i < lines.length; i++) {
      var cells = lines[i].split('\t');
      wide = Math.max(wide, cells.length);
      for (var j = 0; j < cells.length; j++) {
        var td = cellAt(cur.r + i, cur.c + j);
        if (td && td.textContent !== cells[j]) writes.push({td: td, text: cells[j]});
      }
    }
    var r0 = cur.r, c0 = cur.c;
    writeCells(writes);
    anchor = {r: r0, c: c0};
    moveTo(Math.min(tbody.rows.length - 1, r0 + lines.length - 1),
           Math.min(colCount(), c0 + wide - 1), true);
  }
  function clearSelection(){
    if (!editable) return;
    var s = selRange(), writes = [];
    for (var i = s.r1; i <= s.r2; i++) {
      for (var j = s.c1; j <= s.c2; j++) {
        var td = cellAt(i, j);
        if (td && td.textContent !== '') writes.push({td: td, text: ''});
      }
    }
    writeCells(writes);
  }

  /* ---------- structural edits (mirror what openpyxl will do) ---------- */
  function renumberRows(){
    for (var i = 0; i < tbody.rows.length; i++) {
      var tr = tbody.rows[i], n = rowOffset + i + 1;
      tr.dataset.r = String(n);
      var th = tr.cells[0];
      th.textContent = String(n);
      var grip = document.createElement('span');
      grip.className = 'dsheet-grip-row';
      th.appendChild(grip);
    }
  }
  function relabelCols(){
    for (var i = 1; i < thead.cells.length; i++) {
      var th = thead.cells[i];
      th.textContent = colName(i);
      var grip = document.createElement('span');
      grip.className = 'dsheet-grip';
      th.appendChild(grip);
    }
  }
  function insertRows(at, count){
    if (!editable || !structuralAllowed()) return;
    var idx = at - 1 - rowOffset;
    if (idx < 0 || idx > tbody.rows.length) return;
    var width = colCount();
    for (var k = 0; k < count; k++) {
      var tr = tbody.insertRow(idx);
      tr.appendChild(document.createElement('th'));
      for (var j = 0; j < width; j++) tr.insertCell(-1);
    }
    renumberRows();
    push([{op: 'insert_rows', at: at, count: count}]);
    remember({
      undo: function(){ deleteRows(at, count); },
      redo: function(){ insertRows(at, count); }
    });
    moveTo(idx, cur.c, false);
  }
  function deleteRows(at, count){
    if (!editable || !structuralAllowed()) return;
    var idx = at - 1 - rowOffset;
    if (idx < 0 || idx >= tbody.rows.length) return;
    count = Math.min(count, tbody.rows.length - idx);
    var removed = [];
    for (var k = 0; k < count; k++) {
      var tr = tbody.rows[idx + k], values = [];
      for (var j = 1; j < tr.cells.length; j++) values.push(cellText(tr.cells[j]));
      removed.push(values);
    }
    for (var k2 = 0; k2 < count; k2++) tbody.deleteRow(idx);
    renumberRows();
    push([{op: 'delete_rows', at: at, count: count}]);
    remember({
      undo: function(){
        insertRows(at, count);
        var writes = [];
        for (var r = 0; r < removed.length; r++) {
          for (var c = 0; c < removed[r].length; c++) {
            if (removed[r][c] === '') continue;
            var td = cellAt(at - 1 - rowOffset + r, c + 1);
            if (td) writes.push({td: td, text: removed[r][c]});
          }
        }
        writeCells(writes);
      },
      redo: function(){ deleteRows(at, count); }
    });
    moveTo(Math.min(idx, tbody.rows.length - 1), cur.c, false);
  }
  function insertCols(at, count){
    if (!editable || !structuralAllowed()) return;
    if (at < 1 || at > colCount() + 1) return;
    for (var k = 0; k < count; k++) {
      thead.insertBefore(document.createElement('th'), thead.cells[at] || null);
      var col = document.createElement('col');
      col.style.width = '110px';
      colgroup.insertBefore(col, colgroup.querySelectorAll('col')[at] || null);
      for (var i = 0; i < tbody.rows.length; i++) {
        var tr = tbody.rows[i];
        tr.insertBefore(document.createElement('td'), tr.cells[at] || null);
      }
    }
    relabelCols();
    retotal();
    push([{op: 'insert_cols', at: at, count: count}]);
    remember({
      undo: function(){ deleteCols(at, count); },
      redo: function(){ insertCols(at, count); }
    });
    moveTo(cur.r, at, false);
  }
  function deleteCols(at, count){
    if (!editable || !structuralAllowed()) return;
    if (at < 1 || at > colCount()) return;
    count = Math.min(count, colCount() - at + 1);
    var removedCols = [];
    for (var i0 = 0; i0 < tbody.rows.length; i0++) {
      var row = [];
      for (var k0 = 0; k0 < count; k0++) {
        var cell = tbody.rows[i0].cells[at + k0];
        row.push(cell ? cellText(cell) : '');
      }
      removedCols.push(row);
    }
    for (var k = 0; k < count; k++) {
      thead.deleteCell(at);
      var cols = colgroup.querySelectorAll('col');
      if (cols[at]) colgroup.removeChild(cols[at]);
      for (var i = 0; i < tbody.rows.length; i++) tbody.rows[i].deleteCell(at);
    }
    relabelCols();
    retotal();
    push([{op: 'delete_cols', at: at, count: count}]);
    remember({
      undo: function(){
        insertCols(at, count);
        var writes = [];
        for (var r = 0; r < removedCols.length; r++) {
          for (var c = 0; c < removedCols[r].length; c++) {
            if (removedCols[r][c] === '') continue;
            var td = cellAt(r, at + c);
            if (td) writes.push({td: td, text: removedCols[r][c]});
          }
        }
        writeCells(writes);
      },
      redo: function(){ deleteCols(at, count); }
    });
    moveTo(cur.r, Math.min(at, colCount()), false);
  }

  /* ---------- sort / filter (view only, never written to the file) ---------- */
  function applySort(c){
    var rows = Array.prototype.slice.call(tbody.rows);
    if (sortState && sortState.col === c && sortState.dir === -1) {
      rows.sort(function(a, b){ return (a._dsheetHome || 0) - (b._dsheetHome || 0); });
      sortState = null;
    } else {
      var dir = (sortState && sortState.col === c && sortState.dir === 1) ? -1 : 1;
      sortState = {col: c, dir: dir};
      rows.forEach(function(tr, i){ if (tr._dsheetHome === undefined) tr._dsheetHome = i; });
      rows.sort(function(a, b){
        var x = (a.cells[c] ? a.cells[c].textContent : '').trim();
        var y = (b.cells[c] ? b.cells[c].textContent : '').trim();
        if (x === '' && y !== '') return 1;
        if (y === '' && x !== '') return -1;
        var nx = parseFloat(x), ny = parseFloat(y);
        if (!isNaN(nx) && !isNaN(ny) && /^[-+0-9.eE]+$/.test(x) && /^[-+0-9.eE]+$/.test(y)) {
          return (nx - ny) * dir;
        }
        return x.localeCompare(y, undefined, {numeric: true, sensitivity: 'base'}) * dir;
      });
    }
    for (var i = 0; i < rows.length; i++) tbody.appendChild(rows[i]);
    flash(sortState
      ? ('Sorted by column ' + colName(c) + (sortState.dir === 1 ? ' ↑' : ' ↓') + ' - view only')
      : 'Sort cleared');
  }
  function applyFilter(q){
    var needle = String(q || '').trim().toLowerCase();
    filterActive = needle !== '';
    for (var i = 0; i < tbody.rows.length; i++) {
      var tr = tbody.rows[i];
      if (!needle) { tr.style.display = ''; continue; }
      var hit = false;
      for (var j = 1; j < tr.cells.length && !hit; j++) {
        if ((tr.cells[j].textContent || '').toLowerCase().indexOf(needle) >= 0) hit = true;
      }
      tr.style.display = hit ? '' : 'none';
    }
  }

  /* ---------- context menu ---------- */
  function closeMenu(){
    if (menu && menu.parentNode) menu.parentNode.removeChild(menu);
    menu = null;
  }
  function openMenu(x, y, items){
    closeMenu();
    menu = document.createElement('div');
    menu.className = 'dsheet-menu';
    items.forEach(function(it){
      var d = document.createElement('div');
      d.textContent = it.label;
      d.addEventListener('click', function(ev){ ev.stopPropagation(); closeMenu(); it.run(); });
      menu.appendChild(d);
    });
    menu.style.left = Math.min(x, window.innerWidth - 210) + 'px';
    menu.style.top = Math.min(y, window.innerHeight - 190) + 'px';
    document.body.appendChild(menu);
  }
  document.addEventListener('mousedown', function(e){
    if (menu && !menu.contains(e.target)) closeMenu();
  }, true);

  /* ---------- events ---------- */
  var dragging = false;

  scroll.addEventListener('mousedown', function(e){
    var t = e.target;
    if (t.classList && t.classList.contains('dsheet-grip')) { startColResize(e); return; }
    if (t.classList && t.classList.contains('dsheet-grip-row')) { startRowResize(e); return; }
    var td = t.closest ? t.closest('td') : null;
    var th = t.closest ? t.closest('th') : null;
    if (editing && editing.td === td) return;
    if (editing) commitEdit();
    if (th && th.parentNode === thead && th.cellIndex > 0) {
      /* Whole column. With the office feel, anchor at the far end so the
         active cell is the top one and the viewport holds still: revealing
         the last row would drop the user at the bottom of the sheet for
         wanting to pick a column. */
      e.preventDefault();
      if (OFFICE) {
        anchor = {r: tbody.rows.length - 1, c: th.cellIndex};
        moveTo(0, th.cellIndex, true, true);
      } else {
        anchor = {r: 0, c: th.cellIndex};
        moveTo(tbody.rows.length - 1, th.cellIndex, true);
      }
      scroll.focus({preventScroll: true});
      return;
    }
    if (th && th.parentNode.parentNode === tbody) {
      /* Whole row, the same reasoning sideways. */
      e.preventDefault();
      var ri = rowIndexOf(th.parentNode);
      if (OFFICE) {
        anchor = {r: ri, c: colCount()};
        moveTo(ri, 1, true, true);
      } else {
        anchor = {r: ri, c: 1};
        moveTo(ri, colCount(), true);
      }
      scroll.focus({preventScroll: true});
      return;
    }
    if (!td || td.parentNode.parentNode !== tbody) return;
    e.preventDefault();
    var r = rowIndexOf(td.parentNode);
    if (e.shiftKey) { moveTo(r, td.cellIndex, true); }
    else { anchor = {r: r, c: td.cellIndex}; moveTo(r, td.cellIndex, false); dragging = true; }
    scroll.focus({preventScroll: true});
  }, false);

  scroll.addEventListener('mouseover', function(e){
    if (!dragging) return;
    var td = e.target.closest ? e.target.closest('td') : null;
    if (!td || td.parentNode.parentNode !== tbody) return;
    moveTo(rowIndexOf(td.parentNode), td.cellIndex, true);
  }, false);

  document.addEventListener('mouseup', function(){ dragging = false; });

  scroll.addEventListener('dblclick', function(e){
    var td = e.target.closest ? e.target.closest('td') : null;
    if (td && td.parentNode.parentNode === tbody) { e.preventDefault(); beginEdit(td, null); return; }
    var th = e.target.closest ? e.target.closest('th') : null;
    if (th && th.parentNode === thead && th.cellIndex > 0) { e.preventDefault(); applySort(th.cellIndex); }
  }, false);

  scroll.addEventListener('contextmenu', function(e){
    if (!editable) return;
    var td = e.target.closest ? e.target.closest('td') : null;
    var th = e.target.closest ? e.target.closest('th') : null;
    e.preventDefault();
    var rowNo = null, colNo = null;
    if (td && td.parentNode.parentNode === tbody) {
      rowNo = parseInt(td.parentNode.dataset.r, 10); colNo = td.cellIndex;
    } else if (th && th.parentNode === thead && th.cellIndex > 0) {
      colNo = th.cellIndex;
    } else if (th && th.parentNode.parentNode === tbody) {
      rowNo = parseInt(th.parentNode.dataset.r, 10);
    }
    var items = [];
    if (rowNo !== null) {
      items.push({label: 'Insert row above', run: function(){ insertRows(rowNo, 1); }});
      items.push({label: 'Insert row below', run: function(){ insertRows(rowNo + 1, 1); }});
      items.push({label: 'Delete row', run: function(){ deleteRows(rowNo, 1); }});
    }
    if (colNo !== null) {
      items.push({label: 'Insert column left', run: function(){ insertCols(colNo, 1); }});
      items.push({label: 'Insert column right', run: function(){ insertCols(colNo + 1, 1); }});
      items.push({label: 'Delete column', run: function(){ deleteCols(colNo, 1); }});
      if (editable && wrap.dataset.kind === 'xlsx') {
        /* These reorder the sheet itself, over every row of it and not
           only the ones on screen. Clicking a heading still sorts the view
           and changes nothing -- that is the one to reach for by accident. */
        items.push({label: 'Sort A → Z (whole sheet)',
                    run: function(){ send('sort', [], {col: colNo}); }});
        items.push({label: 'Sort Z → A (whole sheet)',
                    run: function(){ send('sort', [], {col: colNo, desc: true}); }});
      }
    }
    if (items.length) openMenu(e.clientX, e.clientY, items);
  }, false);

  scroll.addEventListener('keydown', function(e){
    if (editing) {
      if (e.key === 'Escape') { e.preventDefault(); e.stopPropagation(); cancelEdit(); }
      else if (e.key === 'Enter' && !e.shiftKey && !e.altKey) {
        e.preventDefault(); e.stopPropagation();
        commitEdit();
        moveTo(cur.r + 1, cur.c, false);
      } else if (e.key === 'Tab') {
        e.preventDefault(); e.stopPropagation();
        var back = e.shiftKey;
        commitEdit();
        moveTo(cur.r, cur.c + (back ? -1 : 1), false);
      }
      return;
    }
    var handled = true;
    var ctrl = e.ctrlKey || e.metaKey;
    if (ctrl && (e.key === 'c' || e.key === 'C' || e.key === 'x' || e.key === 'X')) {
      /* Let the native copy/cut event fire -- it works over plain http and
         needs no clipboard permission. Fall back if the browser stays quiet. */
      copyHandled = false;
      var tsv = selToTsv();
      var cut = (e.key === 'x' || e.key === 'X');
      setTimeout(function(){
        if (!copyHandled) copyFallback(tsv);
        if (cut) clearSelection();
      }, 30);
      handled = false;
    }
    else if (ctrl && (e.key === 'v' || e.key === 'V')) { handled = false; }
    else if (ctrl && (e.key === 's' || e.key === 'S')) { if (pending) send('save'); }
    else if (ctrl && (e.key === 'b' || e.key === 'B')) {
      applyFormat({bold: !selectionHas('bold')});
    }
    else if (ctrl && (e.key === 'i' || e.key === 'I')) {
      applyFormat({italic: !selectionHas('italic')});
    }
    else if (ctrl && (e.key === 'u' || e.key === 'U')) {
      applyFormat({underline: !selectionHas('underline')});
    }
    else if (ctrl && (e.key === 'z' || e.key === 'Z')) { e.shiftKey ? redo() : undo(); }
    else if (ctrl && (e.key === 'y' || e.key === 'Y')) { redo(); }
    else if (ctrl && (e.key === 'a' || e.key === 'A')) {
      if (OFFICE) { anchor = {r: tbody.rows.length - 1, c: colCount()};
                    moveTo(0, 1, true, true); }
      else { anchor = {r: 0, c: 1};
             moveTo(tbody.rows.length - 1, colCount(), true); }
    }
    else if (e.key === 'ArrowUp') { moveTo(cur.r - 1, cur.c, e.shiftKey); }
    else if (e.key === 'ArrowDown') { moveTo(cur.r + 1, cur.c, e.shiftKey); }
    else if (e.key === 'ArrowLeft') { moveTo(cur.r, cur.c - 1, e.shiftKey); }
    else if (e.key === 'ArrowRight') { moveTo(cur.r, cur.c + 1, e.shiftKey); }
    else if (e.key === 'PageDown') { moveTo(cur.r + 25, cur.c, e.shiftKey); }
    else if (e.key === 'PageUp') { moveTo(cur.r - 25, cur.c, e.shiftKey); }
    else if (e.key === 'Home') { moveTo(ctrl ? 0 : cur.r, 1, e.shiftKey); }
    else if (e.key === 'End') { moveTo(ctrl ? tbody.rows.length - 1 : cur.r, colCount(), e.shiftKey); }
    else if (e.key === 'Tab') { moveTo(cur.r, cur.c + (e.shiftKey ? -1 : 1), false); }
    else if (e.key === 'Enter' || e.key === 'F2') { beginEdit(cellAt(cur.r, cur.c), null); }
    else if (e.key === 'Delete' || e.key === 'Backspace') { clearSelection(); }
    else if (!ctrl && !e.altKey && e.key.length === 1) { beginEdit(cellAt(cur.r, cur.c), e.key); }
    else { handled = false; }
    if (handled) { e.preventDefault(); e.stopPropagation(); }
  }, false);

  scroll.addEventListener('copy', function(e){
    if (editing) return;
    copyHandled = true;
    e.preventDefault();
    e.stopPropagation();
    if (e.clipboardData) e.clipboardData.setData('text/plain', selToTsv());
  }, false);

  scroll.addEventListener('cut', function(e){
    if (editing) return;
    copyHandled = true;
    e.preventDefault();
    e.stopPropagation();
    if (e.clipboardData) e.clipboardData.setData('text/plain', selToTsv());
  }, false);

  scroll.addEventListener('paste', function(e){
    e.preventDefault();
    e.stopPropagation();
    var text = (e.clipboardData || window.clipboardData).getData('text');
    if (editing) {
      /* plain text only -- pasting into a contenteditable cell otherwise
         injects the source document's markup */
      document.execCommand('insertText', false, String(text).replace(/[\r\n\t]/g, ' '));
      return;
    }
    pasteTsv(text);
  }, false);

  scroll.addEventListener('focusout', function(e){
    if (editing && e.target === editing.td) commitEdit();
  }, false);

  /* ---------- resizing ---------- */
  function startColResize(e){
    e.preventDefault();
    e.stopPropagation();
    var idx = e.target.parentNode.cellIndex;
    var col = colgroup.querySelectorAll('col')[idx];
    if (!col) return;
    var startX = e.clientX, startW = parseInt(col.style.width, 10) || 110;
    function onMove(ev){
      col.style.width = Math.max(34, startW + (ev.clientX - startX)) + 'px';
      retotal();
    }
    function onUp(){
      document.removeEventListener('mousemove', onMove, true);
      document.removeEventListener('mouseup', onUp, true);
    }
    document.addEventListener('mousemove', onMove, true);
    document.addEventListener('mouseup', onUp, true);
  }
  function startRowResize(e){
    e.preventDefault();
    e.stopPropagation();
    var tr = e.target.parentNode.parentNode;
    var startY = e.clientY, startH = tr.getBoundingClientRect().height;
    function onMove(ev){
      var h = Math.max(16, startH + (ev.clientY - startY));
      for (var i = 0; i < tr.cells.length; i++) tr.cells[i].style.height = h + 'px';
    }
    function onUp(){
      document.removeEventListener('mousemove', onMove, true);
      document.removeEventListener('mouseup', onUp, true);
    }
    document.addEventListener('mousemove', onMove, true);
    document.addEventListener('mouseup', onUp, true);
  }

  /* ---------- toolbar ---------- */
  if (undoBtn) undoBtn.addEventListener('click', undo);
  if (redoBtn) redoBtn.addEventListener('click', redo);
  if (saveBtn) saveBtn.addEventListener('click', function(){ if (pending) send('save'); });
  if (discardBtn) discardBtn.addEventListener('click', function(){ send('discard'); });
  if (filterEl) {
    filterEl.addEventListener('input', function(){ applyFilter(filterEl.value); });
    filterEl.addEventListener('keydown', function(ev){ ev.stopPropagation(); });
  }
  Array.prototype.forEach.call(wrap.querySelectorAll('.dsheet-bar .dsheet-btn[data-act]'), function(btn){
    btn.addEventListener('click', function(){
      var tr = tbody.rows[cur.r];
      var rowNo = tr ? parseInt(tr.dataset.r, 10) : rowOffset + 1;
      var act = btn.getAttribute('data-act');
      if (act === 'insert_rows') insertRows(rowNo, 1);
      else if (act === 'delete_rows') deleteRows(rowNo, 1);
      else if (act === 'insert_cols') insertCols(cur.c, 1);
      else if (act === 'delete_cols') deleteCols(cur.c, 1);
    });
  });
  var tabsBar = wrap.querySelector('.dsheet-tabs');
  if (tabsBar) {
    var addTab = tabsBar.querySelector('.dsheet-tab-add');
    var renameTab = tabsBar.querySelector('.dsheet-tab-rename');
    var dropTab = tabsBar.querySelector('.dsheet-tab-drop');

    /* Asked for in the tab strip rather than in a dialog: the name is being
       typed where the name will appear. */
    function askForName(current, action){
      if (tabsBar.querySelector('.dsheet-name-input')) return;
      var box = document.createElement('input');
      box.className = 'dsheet-name-input';
      box.value = current || '';
      box.placeholder = 'Sheet name';
      tabsBar.appendChild(box);
      box.focus();
      box.select();
      var done = false;
      function finish(ok){
        if (done) return;
        done = true;
        var name = box.value.trim();
        box.remove();
        if (ok && name) send(action, [], {name: name});
      }
      box.addEventListener('keydown', function(e){
        if (e.key === 'Enter') { e.preventDefault(); finish(true); }
        else if (e.key === 'Escape') { e.preventDefault(); finish(false); }
        e.stopPropagation();
      });
      box.addEventListener('blur', function(){ finish(false); });
    }

    if (addTab) addTab.addEventListener('click', function(){
      askForName('', 'new_sheet');
    });
    if (renameTab) renameTab.addEventListener('click', function(){
      askForName(wrap.dataset.sheet || '', 'rename_sheet');
    });
    if (dropTab) dropTab.addEventListener('click', function(){
      /* Two clicks, because a sheet is not something to lose to a slip. */
      if (dropTab.dataset.armed === '1') {
        dropTab.dataset.armed = '0';
        dropTab.textContent = '\u2212';
        send('drop_sheet', []);
        return;
      }
      dropTab.dataset.armed = '1';
      dropTab.textContent = 'Delete?';
      setTimeout(function(){
        if (dropTab.dataset.armed !== '1') return;
        dropTab.dataset.armed = '0';
        dropTab.textContent = '\u2212';
      }, 4000);
    });
  }

  Array.prototype.forEach.call(
      wrap.querySelectorAll('.dsheet-tab[data-sheet]'), function(tab){
    tab.addEventListener('click', function(){
      if (tab.classList.contains('dsheet-tab-on')) return;
      send('switch_sheet', [], {target: tab.getAttribute('data-sheet')});
    });
  });
  Array.prototype.forEach.call(wrap.querySelectorAll('.dsheet-foot .dsheet-btn[data-page]'), function(btn){
    btn.addEventListener('click', function(){
      send('page', [], {dir: btn.getAttribute('data-page')});
    });
  });

  /* ---------- init ---------- */
  reflectPending();
  reflectHistory();
  /* Seed the cursor without revealing it: the saved scroll position is set
     right after, and a reveal here would first drag the sheet back to A1.
     data-cursor is where the user was when something forced a rebuild. */
  var seed = (wrap.dataset.cursor || '0,1').split(',');
  moveTo(parseInt(seed[0], 10) || 0, parseInt(seed[1], 10) || 1, false, true);
  var st = parseInt(wrap.dataset.scrolltop || '0', 10) || 0;
  if (st) scroll.scrollTop = st;
  scroll.focus({preventScroll: true});
  }

  boot();
})();
"""


def grid_js(scope_class: str, token: str) -> str:
    """JavaScript that makes a rendered grid interactive.

    Must be injected through the dashboard's ``run_js`` channel: ipywidgets sets
    HTML widget content via ``innerHTML``, which never executes inline scripts.
    The script retries until its grid appears, and binds once per rendered grid.
    """
    return (
        _GRID_JS_TEMPLATE
        .replace('__SCOPE__', json.dumps(str(scope_class)))
        .replace('__TOKEN__', json.dumps(str(token)))
    )
