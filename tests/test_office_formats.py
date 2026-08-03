"""The four file kinds that used to stop an office run: ODF, .xls, scans, XFA.

Every case here is one where a wrong answer beats no answer only in
appearance. An administrative file read through an approximate parser
comes back as rows that look exactly like data, and nothing downstream
can tell that they are not what the file says. So what is pinned is as
much what the module REFUSES as what it reads:

* an .ods whose repeated cells are expanded, so the values stay under
  their own headings, and whose stored numbers are read instead of their
  German rendering;
* a legacy .xls refused with the command that converts it, and an HTML
  export wearing an .xls name called what it actually is;
* a page with no text layer told apart from an empty page, with the
  missing OCR component named rather than the capability;
* an XFA form whose values can be read out of its XML dataset, and a
  dynamic one refused for filling instead of reported as filled.

All fixtures are built here in code — there are no sample documents in
the repository.
"""

from __future__ import annotations

import datetime
import json
from pathlib import Path

import pytest

from delfin.agent import office
from delfin.agent.api_client import (
    _DOC_TOOLS_OPENAI,
    _DocToolExecutor,
    _OFFICE_TOOL_BACKENDS,
    KitToolPermissions,
    _binary_read_hint,
)

odf = pytest.importorskip("odf")
pypdf = pytest.importorskip("pypdf")


# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------

@pytest.fixture
def ws(tmp_path):
    d = tmp_path / "ws"
    d.mkdir()
    return d


def _perms(root: Path) -> KitToolPermissions:
    perms = KitToolPermissions(workspace=str(root))
    perms.mode = "acceptEdits"
    perms.task_session_id = "office-formats-test"
    return perms


def _cell(value=None, *, repeat=0, formula="", covered=False, hidden=False):
    """One .ods cell, typed the way LibreOffice types it."""
    from odf.table import CoveredTableCell, TableCell
    from odf.text import P

    if covered:
        return CoveredTableCell()
    attrs: dict = {}
    text = ""
    if isinstance(value, bool):
        attrs.update(valuetype="boolean", booleanvalue=str(value).lower())
        text = "WAHR" if value else "FALSCH"
    elif isinstance(value, (int, float)):
        attrs.update(valuetype="float", value=value)
        # The display text is the German rendering; the stored value is
        # the number. A reader that takes the text back is the whole
        # decimal-comma problem again.
        text = (f"{value:,.2f}".replace(",", "#")
                .replace(".", ",").replace("#", "."))
    elif isinstance(value, datetime.date):
        attrs.update(valuetype="date", datevalue=value.isoformat())
        text = value.strftime("%d.%m.%Y")
    elif value is not None:
        attrs.update(valuetype="string")
        text = str(value)
    if formula:
        attrs["formula"] = formula
    if repeat:
        attrs["numbercolumnsrepeated"] = repeat
    cell = TableCell(**attrs)
    if text:
        cell.addElement(P(text=text))
    return cell


def _ods(path: Path, sheets: dict, *, trailing_filler=True, hidden_rows=()):
    """Write an .ods from ``{sheet name: [[cell, ...], ...]}``."""
    from odf.opendocument import OpenDocumentSpreadsheet
    from odf.table import Table, TableRow

    document = OpenDocumentSpreadsheet()
    for name, rows in sheets.items():
        table = Table(name=name)
        for index, row in enumerate(rows, start=1):
            attrs = ({"visibility": "collapse"} if index in hidden_rows else {})
            element = TableRow(**attrs)
            for cell in row:
                element.addElement(cell)
            table.addElement(element)
        if trailing_filler:
            # What LibreOffice writes for the rest of the sheet. A reader
            # that counts it reports a million rows for a four-row file.
            table.addElement(TableRow(numberrowsrepeated=1048571))
        document.spreadsheet.addElement(table)
    document.save(str(path))
    return path


@pytest.fixture
def buch(ws):
    """A small ledger with the shapes an .ods really carries."""
    return _ods(ws / "buch.ods", {
        "Daten": [
            [_cell("Belegnr"), _cell("Betrag"), _cell("Datum"),
             _cell(repeat=1010)],
            [_cell("R-001"), _cell(1234.5), _cell(datetime.date(2026, 7, 31))],
            [_cell("R-002"), _cell(-7.25), _cell(datetime.date(2026, 1, 3))],
            [_cell("Summe"),
             _cell(1227.25, formula="of:=SUM([.B2:.B3])"),
             _cell(covered=True)],
        ],
        "Notizen": [[_cell("nur eine Notiz")]],
    })


def _odt(path: Path):
    from odf.opendocument import OpenDocumentText
    from odf.table import Table, TableCell, TableRow
    from odf.text import H, List, ListItem, P

    document = OpenDocumentText()
    document.text.addElement(H(outlinelevel=1, text="Bescheid"))
    document.text.addElement(P(text="Sehr geehrte Frau Müller,"))
    items = List()
    entry = ListItem()
    entry.addElement(P(text="Anlage: Kostenaufstellung"))
    items.addElement(entry)
    document.text.addElement(items)
    table = Table(name="Posten")
    for values in (("Posten", "Betrag"), ("Reise", "240,00")):
        row = TableRow()
        for value in values:
            cell = TableCell()
            cell.addElement(P(text=value))
            row.addElement(cell)
        table.addElement(row)
    document.text.addElement(table)
    document.save(str(path))
    return path


# ---------------------------------------------------------------------------
# OpenDocument spreadsheets
# ---------------------------------------------------------------------------

def test_ods_is_read_as_an_addressable_grid(buch):
    result = office.read_sheet(buch)
    assert result["kind"] == "spreadsheet"
    assert result["format"] == "ods"
    assert result["writable"] is False
    assert result["sheets"] == ["Daten", "Notizen"]
    assert result["sheet"] == "Daten"
    assert "R-001" in result["grid"] and "B" in result["grid"]


def test_repeated_cells_and_rows_do_not_invent_a_sheet(buch):
    """number-columns/rows-repeated is filler, not content."""
    result = office.read_sheet(buch)
    # Four rows and three columns — not 1 048 575 rows and 1 013 columns.
    assert result["rows"] == 4
    assert result["columns"] == 3


def test_a_repeat_in_the_middle_of_a_row_keeps_the_columns_aligned(ws):
    """The failure this guards: a value sliding into the wrong column."""
    path = _ods(ws / "luecke.ods", {"Daten": [
        [_cell("A"), _cell(repeat=3), _cell("E")],
        [_cell("a2"), _cell(repeat=3), _cell("e2")],
    ]}, trailing_filler=False)
    rows = office._table_rows(path)
    assert rows[0] == ["A", "", "", "", "E"]
    assert rows[1][4] == "e2"


def test_stored_numbers_are_read_not_their_german_rendering(buch):
    """The cell shows 1.234,50; the file stores 1234.5. Take the number."""
    rows = office._table_rows(buch)
    assert rows[1][1] == 1234.5
    assert rows[2][1] == -7.25
    profile = {p["name"]: p for p in office.read_sheet(buch)["column_profile"]}
    assert profile["Datum"]["kind"] == "date"
    assert profile["Datum"]["convention"] == office.ISO_DATE
    assert rows[1][2] == datetime.date(2026, 7, 31)


def test_a_formula_without_a_stored_result_is_shown_as_a_formula(ws):
    path = _ods(ws / "offen.ods", {"Daten": [
        [_cell("Posten"), _cell("Betrag")],
        [_cell("Summe"), _cell(formula="of:=SUM([.B1:.B1])")],
    ]}, trailing_filler=False)
    result = office.read_sheet(path)
    assert "=SUM" in result["grid"]
    assert any("no stored result" in n for n in result["notes"])


def test_a_cached_formula_result_is_used(buch):
    result = office.read_sheet(buch)
    assert "1227.25" in result["grid"]
    assert not any("no stored result" in n for n in result["notes"])


def test_hidden_rows_are_reported_rather_than_passed_off_as_the_sheet(ws):
    path = _ods(ws / "gefiltert.ods", {"Daten": [
        [_cell("Name"), _cell("Betrag")],
        [_cell("Meier"), _cell(100.0)],
        [_cell("Schulz"), _cell(200.0)],
    ]}, trailing_filler=False, hidden_rows=(3,))
    result = office.read_sheet(path)
    assert any("hidden" in n for n in result["notes"])


def test_covered_cells_of_a_merge_keep_their_position(ws):
    path = _ods(ws / "merge.ods", {"Daten": [
        [_cell("A"), _cell(covered=True), _cell("C")],
    ]}, trailing_filler=False)
    assert office._table_rows(path)[0] == ["A", None, "C"]


def test_a_named_sheet_can_be_chosen_and_an_unknown_one_lists_the_names(buch):
    assert office.read_sheet(buch, sheet="Notizen")["sheet"] == "Notizen"
    with pytest.raises(office.OfficeError) as exc:
        office.read_sheet(buch, sheet="Quartal")
    assert "Notizen" in str(exc.value)


def test_paging_an_ods_reports_the_window_and_the_total(ws):
    path = _ods(ws / "lang.ods", {"Daten": [
        [_cell(f"Zeile {n}")] for n in range(1, 21)
    ]}, trailing_filler=False)
    result = office.read_sheet(path, start_row=5, max_rows=3)
    assert result["rows"] == 20
    assert "Zeile 5" in result["grid"] and "Zeile 8" not in result["grid"]
    assert any("start_row" in n for n in result["notes"])


def test_an_ods_reaches_compare_tables(ws, buch):
    """The point of the reader: reconciling an .ods against another export."""
    other = ws / "kasse.csv"
    other.write_text(
        "Belegnr;Betrag\nR-001;1.234,50\nR-002;-7,25\nR-003;5,00\n",
        encoding="utf-8")
    result = office.compare_tables(buch, other, key="Belegnr",
                                   columns=["Betrag"])
    # 1.234,50 out of the CSV equals the 1234.5 stored in the .ods.
    assert set(result["equal"]) == {"R-001", "R-002"}
    assert result["only_right"] == ["R-003"]
    assert result["rows_accounted_for"] is True
    assert any("ACTIVE sheet" in n for n in result["notes"])


def test_writing_an_ods_is_refused_and_names_the_conversion(buch):
    for call in (
        lambda: office.edit_sheet(buch, edits=[{"cell": "A1", "value": "x"}]),
        lambda: office.plan_sheet_edits(
            buch, edits=[{"cell": "A1", "value": "x"}]),
    ):
        with pytest.raises(office.OfficeError) as exc:
            call()
        message = str(exc.value)
        assert "read here but not written" in message
        assert "convert-to xlsx" in message


def test_creating_an_ods_is_refused_with_the_same_route(ws):
    with pytest.raises(office.OfficeError) as exc:
        office.create_sheet(ws / "neu.ods", [["a", "b"]])
    assert "xlsx" in str(exc.value)


def test_an_ods_that_is_really_an_xlsx_is_named_for_what_it_is(ws):
    import openpyxl

    path = ws / "getarnt.ods"
    book = openpyxl.Workbook()
    book.active.append(["Name", "Betrag"])
    book.save(path)
    with pytest.raises(office.OfficeError) as exc:
        office.read_sheet(path)
    assert "xlsx" in str(exc.value).lower()


def test_an_odt_named_ods_is_refused_rather_than_read_as_a_table(ws):
    path = _odt(ws / "text.odt")
    renamed = ws / "text.ods"
    renamed.write_bytes(path.read_bytes())
    with pytest.raises(office.OfficeError) as exc:
        office.read_sheet(renamed)
    assert "opendocument.text" in str(exc.value)


def test_a_damaged_ods_says_so_instead_of_returning_nothing(ws):
    path = ws / "kaputt.ods"
    path.write_bytes(b"not a zip at all")
    with pytest.raises(office.OfficeError) as exc:
        office.read_sheet(path)
    assert "kaputt.ods" in str(exc.value)


# ---------------------------------------------------------------------------
# OpenDocument text
# ---------------------------------------------------------------------------

def test_odt_paragraphs_tables_and_lists_are_all_extracted(ws):
    result = office.read_document(_odt(ws / "bescheid.odt"))
    assert result["kind"] == "opendocument_text"
    assert "Bescheid" in result["text"]
    assert "Müller" in result["text"]
    # A letter puts content in lists and tables; taking only top-level
    # paragraphs returns a document with pieces missing and no sign of it.
    assert "Kostenaufstellung" in result["text"]
    assert "240,00" in result["text"]
    assert result["tables"] == 1


def test_odt_is_marked_read_only(ws):
    result = office.read_odt(_odt(ws / "bescheid.odt"))
    assert result["writable"] is False
    assert any("docx" in n for n in result["notes"])


def test_odt_as_a_template_is_refused_with_the_route(ws):
    path = _odt(ws / "vorlage.odt")
    with pytest.raises(office.OfficeError) as exc:
        office.read_document(path, fields=True)
    assert "convert-to docx" in str(exc.value)

    table = ws / "empfaenger.csv"
    table.write_text("Name\nMeier\n", encoding="utf-8")
    with pytest.raises(office.OfficeError) as exc:
        office.fill_series(table, path, output_dir=ws / "out")
    assert "convert-to docx" in str(exc.value)
    assert not (ws / "out").exists()


# ---------------------------------------------------------------------------
# Legacy .xls
# ---------------------------------------------------------------------------

def _ole2(path: Path) -> Path:
    """A file carrying the OLE2 signature a legacy workbook starts with."""
    path.write_bytes(b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1" + b"\x00" * 64)
    return path


def test_a_legacy_xls_is_refused_with_the_command_that_converts_it(ws):
    path = _ole2(ws / "haushalt.xls")
    for call in (
        lambda: office.read_document(path),
        lambda: office.read_sheet(path),
        lambda: office._table_rows(path),
    ):
        with pytest.raises(office.OfficeError) as exc:
            call()
        message = str(exc.value)
        assert "legacy Excel workbook" in message
        assert "convert-to xlsx" in message
        # No half-support anywhere in the wording.
        assert "haushalt.xls" in message


def test_an_html_export_named_xls_is_called_what_it_is(ws):
    """The routine case: a web system exports HTML under an .xls name."""
    path = ws / "auswertung.xls"
    path.write_bytes(
        b"<html><body><table><tr><td>Beleg</td></tr></table></body></html>")
    with pytest.raises(office.OfficeError) as exc:
        office.read_document(path)
    assert "HTML" in str(exc.value)


def test_a_legacy_doc_is_refused_with_the_route_too(ws):
    with pytest.raises(office.OfficeError) as exc:
        office.read_document(_ole2(ws / "brief.doc"))
    assert "convert-to docx" in str(exc.value)


def test_a_workbook_named_csv_is_refused_instead_of_decoded(ws):
    """latin-1 decodes every byte, so this never fails on its own."""
    import openpyxl

    path = ws / "liste.csv"
    book = openpyxl.Workbook()
    book.active.append(["Name", "Betrag"])
    book.save(path)
    with pytest.raises(office.OfficeError) as exc:
        office.read_sheet(path)
    assert "noise" in str(exc.value)


def test_container_sniffing_names_the_formats_it_can(ws, buch):
    import openpyxl

    book = openpyxl.Workbook()
    book.save(ws / "echt.xlsx")
    assert office.container_kind(buch) == "ods"
    assert office.container_kind(_odt(ws / "t.odt")) == "odt"
    assert office.container_kind(ws / "echt.xlsx") == "xlsx"
    assert office.container_kind(_ole2(ws / "alt.xls")) == "ole2"
    assert office.container_kind(ws / "fehlt.xls") == ""


# ---------------------------------------------------------------------------
# Scanned pages and OCR
# ---------------------------------------------------------------------------

def _pdf_with_image_page(path: Path) -> Path:
    """A page that draws an image and carries no text — a scan in miniature."""
    from pypdf.generic import DecodedStreamObject, DictionaryObject, NameObject

    writer = pypdf.PdfWriter()
    writer.add_blank_page(width=200, height=200)
    image = DecodedStreamObject()
    image.set_data(b"\x00" * 16)
    image[NameObject("/Type")] = NameObject("/XObject")
    image[NameObject("/Subtype")] = NameObject("/Image")
    image[NameObject("/Width")] = NameObject("/4")
    xobjects = DictionaryObject()
    xobjects[NameObject("/Im0")] = writer._add_object(image)
    resources = DictionaryObject()
    resources[NameObject("/XObject")] = xobjects
    writer.pages[0][NameObject("/Resources")] = resources
    with path.open("wb") as fh:
        writer.write(fh)
    return path


def test_a_scan_is_told_apart_from_an_empty_page(ws):
    scan = office.read_pdf(_pdf_with_image_page(ws / "scan.pdf"))
    note = " ".join(scan["notes"])
    assert "scan" in note
    assert scan["pages_without_text"] == [1]

    blank = pypdf.PdfWriter()
    blank.add_blank_page(width=200, height=200)
    with (ws / "leer.pdf").open("wb") as fh:
        blank.write(fh)
    empty = office.read_pdf(ws / "leer.pdf")
    note = " ".join(empty["notes"])
    # A blank sheet is not a scan, and OCR would find nothing on it.
    assert "simply empty" in note
    assert "which is a scan" not in note


def test_the_ocr_diagnosis_names_the_missing_component(ws, monkeypatch):
    """"OCR unavailable" is not actionable; a package name is."""
    import shutil

    monkeypatch.setattr(shutil, "which", lambda name: None)
    monkeypatch.setenv("EASYOCR_MODULE_PATH", str(ws / "no-models"))
    status = office.ocr_availability()
    assert status["available"] is False
    detail = " ".join(status["detail"])
    # pytesseract imports cleanly and then fails at the first call: the
    # component that is missing is the binary, and that is what is named.
    assert "tesseract" in detail
    assert "PATH" in detail or "not installed" in detail
    assert status["next_step"]

    result = office.read_pdf(_pdf_with_image_page(ws / "scan.pdf"))
    note = " ".join(result["notes"])
    assert "OCR is not possible on this machine" in note
    assert "tesseract" in note


def test_ocr_is_refused_rather_than_faked_when_no_engine_is_there(ws,
                                                                 monkeypatch):
    monkeypatch.setattr(office, "ocr_availability", lambda: {
        "available": False, "engine": "", "renderer": "",
        "detail": ["nothing installed"], "next_step": "install an engine"})
    result = office.read_pdf(_pdf_with_image_page(ws / "scan.pdf"), ocr=True)
    assert result["ocr_pages"] == []
    assert any("nothing installed" in n for n in result["notes"])
    assert result["pages_without_text"] == [1]


def test_ocr_leaves_pages_that_have_real_text_alone(ws):
    reportlab = pytest.importorskip("reportlab")
    from reportlab.pdfgen import canvas

    path = ws / "text.pdf"
    c = canvas.Canvas(str(path))
    c.drawString(60, 780, "Rechnung 4711")
    c.save()
    result = office.read_pdf(path, ocr=True)
    assert "Rechnung 4711" in result["text"]
    assert result["ocr_pages"] == []
    assert any("no requested page is a scan" in n for n in result["notes"])


@pytest.mark.skipif(not office.ocr_availability()["available"],
                    reason="no OCR engine on this machine")
def test_ocr_reads_a_scanned_page_and_labels_it_as_ocr(ws):
    fitz = pytest.importorskip("fitz")
    reportlab = pytest.importorskip("reportlab")
    from reportlab.pdfgen import canvas

    # Deliberately small: OCR cost scales with the page area, and the
    # point of this test is the plumbing, not the recogniser.
    source = ws / "quelle.pdf"
    c = canvas.Canvas(str(source), pagesize=(150, 60))
    c.setFont("Helvetica", 20)
    c.drawString(10, 20, "Beleg 4711")
    c.save()

    rendered = fitz.open(str(source))
    image = rendered[0].get_pixmap(dpi=200).tobytes("png")
    scan = fitz.open()
    page = scan.new_page(width=150, height=60)
    page.insert_image(page.rect, stream=image)
    target = ws / "scan.pdf"
    scan.save(str(target))

    plain = office.read_pdf(target)
    assert plain["pages_without_text"] == [1]

    result = office.read_pdf(target, ocr=True)
    assert "4711" in result["text"].replace(" ", "")
    assert result["ocr_pages"][0]["page"] == 1
    assert 0.0 < result["ocr_pages"][0]["confidence"] <= 1.0
    # The output has to say the text came from pixels, not from the file.
    assert "OCR" in result["text"]
    assert any("reading of the page image" in n for n in result["notes"])


# ---------------------------------------------------------------------------
# XFA forms
# ---------------------------------------------------------------------------

_XFA_DATASETS = (
    b'<?xml version="1.0"?>'
    b'<xfa:datasets xmlns:xfa="http://www.xfa.org/schema/xfa-data/1.0/">'
    b"<xfa:data><Antrag>"
    b"<Antragsteller>M\xc3\xbcller</Antragsteller>"
    b"<Betrag>1234,50</Betrag>"
    b"<Posten><Bezeichnung>Reise</Bezeichnung></Posten>"
    b"<Posten><Bezeichnung>Porto</Bezeichnung></Posten>"
    b"</Antrag></xfa:data></xfa:datasets>"
)


def _xfa_pdf(path: Path, *, dynamic: bool, single_stream: bool = False) -> Path:
    from pypdf.generic import (
        ArrayObject, BooleanObject, DecodedStreamObject, DictionaryObject,
        NameObject, TextStringObject,
    )

    writer = pypdf.PdfWriter()
    writer.add_blank_page(width=200, height=200)
    stream = DecodedStreamObject()
    stream.set_data(_XFA_DATASETS)
    reference = writer._add_object(stream)
    acro = DictionaryObject()
    acro[NameObject("/XFA")] = (
        reference if single_stream
        else ArrayObject([TextStringObject("datasets"), reference]))
    acro[NameObject("/Fields")] = ArrayObject()
    writer._root_object[NameObject("/AcroForm")] = writer._add_object(acro)
    if dynamic:
        writer._root_object[NameObject("/NeedsRendering")] = BooleanObject(True)
    with path.open("wb") as fh:
        writer.write(fh)
    return path


def test_xfa_values_can_be_read_out_of_the_dataset(ws):
    """The listing shows every field empty; the form is not."""
    result = office.pdf_form_fields(_xfa_pdf(ws / "antrag.pdf", dynamic=False))
    assert result["xfa"] is True
    values = {entry["name"]: entry["value"] for entry in result["xfa_values"]}
    assert values["Antrag.Antragsteller"] == "Müller"
    assert values["Antrag.Betrag"] == "1234,50"
    # Repeated siblings keep their position rather than overwriting.
    assert values["Antrag.Posten[1].Bezeichnung"] == "Reise"
    assert values["Antrag.Posten[2].Bezeichnung"] == "Porto"


def test_xfa_values_are_found_in_a_single_stream_form_too(ws):
    path = _xfa_pdf(ws / "einzel.pdf", dynamic=False, single_stream=True)
    assert office.pdf_xfa_data(path)["values"]


def test_a_dynamic_xfa_form_is_recognised_and_refused_for_filling(ws):
    path = _xfa_pdf(ws / "dynamisch.pdf", dynamic=True)
    listing = office.pdf_form_fields(path)
    assert listing["xfa_dynamic"] is True
    assert any("NeedsRendering" in n for n in listing["notes"])

    with pytest.raises(office.OfficeError) as exc:
        office.fill_pdf_form(path, {"irgendwas": "x"}, output=ws / "out.pdf")
    message = str(exc.value)
    assert "Nothing was written" in message
    assert not (ws / "out.pdf").exists()


def test_a_static_xfa_form_is_flagged_but_not_refused(ws):
    """Hybrid forms do render their AcroForm layer — do not over-refuse."""
    path = _xfa_pdf(ws / "statisch.pdf", dynamic=False)
    listing = office.pdf_form_fields(path)
    assert listing["xfa"] is True and listing["xfa_dynamic"] is False
    assert not any("NeedsRendering" in n for n in listing["notes"])


def test_an_xfa_array_that_is_not_name_stream_pairs_does_not_crash(ws):
    from pypdf.generic import (
        ArrayObject, DictionaryObject, NameObject, TextStringObject,
    )

    writer = pypdf.PdfWriter()
    writer.add_blank_page(width=200, height=200)
    acro = DictionaryObject()
    acro[NameObject("/XFA")] = ArrayObject([TextStringObject("<xdp/>")])
    writer._root_object[NameObject("/AcroForm")] = writer._add_object(acro)
    path = ws / "seltsam.pdf"
    with path.open("wb") as fh:
        writer.write(fh)

    listing = office.pdf_form_fields(path)
    assert listing["xfa"] is True
    assert listing["xfa_values"] == []
    assert any("no values could be read" in n.lower() for n in listing["notes"])


def test_read_pdf_reports_that_an_xfa_form_holds_readable_values(ws):
    result = office.read_pdf(_xfa_pdf(ws / "antrag.pdf", dynamic=True))
    note = " ".join(result["notes"])
    assert "dynamic XFA" in note
    assert "fields=true" in note


def test_pdf_xfa_data_refuses_a_document_that_is_not_an_xfa_form(ws):
    writer = pypdf.PdfWriter()
    writer.add_blank_page(width=200, height=200)
    path = ws / "schlicht.pdf"
    with path.open("wb") as fh:
        writer.write(fh)
    with pytest.raises(office.OfficeError):
        office.pdf_xfa_data(path)


# ---------------------------------------------------------------------------
# Wiring: backends, tool surface, doctor
# ---------------------------------------------------------------------------

def test_the_opendocument_backend_is_gated_like_the_others():
    assert "opendocument" in office.available_backends()
    assert office._BACKENDS["opendocument"] == ("odf", "odfpy")
    for tool in ("read_document", "compare_tables"):
        assert "opendocument" in _OFFICE_TOOL_BACKENDS[tool]


def test_a_missing_odfpy_is_reported_as_a_missing_package(ws, buch,
                                                          monkeypatch):
    """Not as a bare ImportError out of the middle of a read."""
    real = office._require

    def without_odf(kind):
        if kind == "opendocument":
            raise office.OfficeError(
                "opendocument support needs the 'odfpy' package")
        return real(kind)

    monkeypatch.setattr(office, "_require", without_odf)
    for call in (lambda: office.read_sheet(buch),
                 lambda: office.read_odt(_odt(ws / "b.odt"))):
        with pytest.raises(office.OfficeError) as exc:
            call()
        assert "odfpy" in str(exc.value)


def test_read_file_points_at_the_reader_for_odf(ws):
    for name in ("haushalt.ods", "bescheid.odt"):
        path = ws / name
        path.write_bytes(b"PK\x03\x04binary")
        hint = _binary_read_hint(path)
        assert hint is not None
        assert "read_document" in hint


def test_the_document_tool_schema_advertises_the_new_formats():
    schema = next(t for t in _DOC_TOOLS_OPENAI
                  if t["function"]["name"] == "read_document")
    description = schema["function"]["description"]
    assert ".ods" in description and ".odt" in description
    assert "ocr" in schema["function"]["parameters"]["properties"]


def test_the_doctor_reports_odf_and_ocr():
    from delfin.agent import doctor

    rows = {r["check"]: r for r in doctor._check_document_backends({})}
    assert "documents: OpenDocument" in rows
    assert "documents: OCR" in rows
    for row in rows.values():
        assert row["status"] in ("PASS", "WARN")


def test_an_ods_reads_through_the_document_tool(ws, buch):
    executor = _DocToolExecutor()
    out = executor._execute_read_document({"path": str(buch)}, _perms(ws))
    assert "R-001" in out
    assert "1234.5" in out


def test_the_xfa_dataset_is_shown_by_the_document_tool(ws):
    path = _xfa_pdf(ws / "antrag.pdf", dynamic=True)
    executor = _DocToolExecutor()
    out = executor._execute_read_document(
        {"path": str(path), "fields": True}, _perms(ws))
    assert "XFA dataset" in out
    assert "Müller" in out


def test_a_legacy_xls_refusal_comes_back_as_an_error_not_a_table(ws):
    path = _ole2(ws / "alt.xls")
    executor = _DocToolExecutor()
    out = executor._execute_read_document({"path": str(path)}, _perms(ws))
    assert "convert-to xlsx" in json.loads(out)["error"]
