"""Paging through a sheet must not re-parse the workbook for every window.

Each window used to cost three streaming passes over the file -- values, cell
formats, formulas -- each of which skips every row above the window.  Moving
down a 5000-row sheet therefore cost the same again on every page.  A sheet that
fits the budget is now read once and sliced.

What matters more than the speed is that the two paths agree: the window the
cache serves has to be the window the file would have given.
"""

from __future__ import annotations

import pytest

from delfin.dashboard import spreadsheet_view as sheet_view

openpyxl = pytest.importorskip("openpyxl")


@pytest.fixture
def workbook(tmp_path):
    """A sheet with everything the two read paths have to agree about."""
    from openpyxl.styles import Font, PatternFill

    book = openpyxl.Workbook()
    ws = book.active
    ws.title = "Data"
    for row in range(1, 1201):
        ws.cell(row=row, column=1, value=f"label {row}")
        ws.cell(row=row, column=2, value=row * 1.5)
        ws.cell(row=row, column=3, value=f"=B{row}*2")
    # formatting that only the style pass can see, inside a later window
    ws.cell(row=700, column=1).font = Font(bold=True, italic=True)
    ws.cell(row=700, column=2).fill = PatternFill("solid", fgColor="FFFF00")
    ws.cell(row=701, column=2).number_format = "0.00"
    path = tmp_path / "book.xlsx"
    book.save(path)
    sheet_view.forget_sheet_cache()
    return path


def _read(path, offset):
    return sheet_view.read_xlsx(path, row_offset=offset)[1]


def test_the_cached_window_is_the_window_the_file_gives(workbook, monkeypatch):
    with monkeypatch.context() as no_cache:      # force the streaming path
        no_cache.setattr(sheet_view, "SHEET_CACHE_MAX_CELLS", 0)
        sheet_view.forget_sheet_cache()
        streamed = _read(workbook, 600)
        assert not sheet_view._SHEET_CACHE

    cached = _read(workbook, 600)            # warm: served from the slice
    assert sheet_view._SHEET_CACHE, "the sheet was never held"

    assert cached.values == streamed.values
    assert cached.styles == streamed.styles
    assert cached.formulas == streamed.formulas
    assert cached.col_widths == streamed.col_widths
    assert cached.row_offset == streamed.row_offset
    assert cached.total_rows == streamed.total_rows
    assert cached.has_formulas == streamed.has_formulas


def test_every_window_agrees_not_just_the_one_that_filled_the_cache(workbook, monkeypatch):
    with monkeypatch.context() as no_cache:
        no_cache.setattr(sheet_view, "SHEET_CACHE_MAX_CELLS", 0)
        sheet_view.forget_sheet_cache()
        expected = {offset: _read(workbook, offset) for offset in (0, 400, 800)}
    sheet_view.forget_sheet_cache()

    _read(workbook, 400)                     # fills the cache
    for offset, want in expected.items():
        got = _read(workbook, offset)
        assert got.values == want.values, f"window at {offset} differs"
        assert got.styles == want.styles, f"formats at {offset} differ"
        assert got.formulas == want.formulas, f"formulas at {offset} differ"


def test_opening_a_file_is_not_slowed_down_by_the_cache(workbook):
    sheet_view.forget_sheet_cache()

    _read(workbook, 0)

    assert not sheet_view._SHEET_CACHE, (
        "reading a whole sheet nobody has scrolled would make opening slower"
    )


def test_a_changed_file_is_not_served_from_memory(workbook):
    _read(workbook, 400)
    _read(workbook, 400)
    assert sheet_view._SHEET_CACHE

    book = openpyxl.load_workbook(workbook)
    book["Data"].cell(row=401, column=1, value="rewritten")
    book.save(workbook)

    fresh = _read(workbook, 400)
    assert fresh.values[0][0] == "rewritten", "the stale window was served"


def test_a_sheet_too_large_to_hold_is_not_held(workbook, monkeypatch):
    monkeypatch.setattr(sheet_view, "SHEET_CACHE_MAX_CELLS", 10)
    sheet_view.forget_sheet_cache()

    _read(workbook, 400)
    _read(workbook, 400)

    assert not sheet_view._SHEET_CACHE, "a sheet over the budget must be streamed"


def test_only_a_couple_of_sheets_are_kept(tmp_path):
    books = []
    for n in range(sheet_view.SHEET_CACHE_ENTRIES + 2):
        book = openpyxl.Workbook()
        ws = book.active
        for row in range(1, 601):
            ws.cell(row=row, column=1, value=f"{n}-{row}")
        path = tmp_path / f"b{n}.xlsx"
        book.save(path)
        books.append(path)
    sheet_view.forget_sheet_cache()

    for path in books:
        _read(path, 400)
        _read(path, 400)

    assert len(sheet_view._SHEET_CACHE) <= sheet_view.SHEET_CACHE_ENTRIES
