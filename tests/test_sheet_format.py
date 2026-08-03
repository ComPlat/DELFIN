"""Bold, italic, a fill colour and a number format on a cell.

What makes a table readable: a heading row that stands out, amounts that
line up, a colour on the rows that belong together. The same controls a
spreadsheet has, doing the same thing, so there is nothing to relearn.

Formatting travels as an op like every other change: it queues with the
edits, saves with them, and one undo takes it off again.
"""

from __future__ import annotations

import json

import pytest

openpyxl = pytest.importorskip('openpyxl')
from openpyxl.styles import Font, PatternFill

from delfin.dashboard import spreadsheet_view as sheet


# ---------------------------------------------------------------------------
# How a value reads
# ---------------------------------------------------------------------------

@pytest.mark.parametrize('code,value,shown', [
    ('0', 1234.5, '1235'),
    ('0.00', 1234.5, '1234,50'),
    ('#,##0.00', 1234.5, '1.234,50'),
    ('0.0%', 0.1234, '12,3 %'),
    ('General', 1234.5, '1234.5'),
    ('@', 5, '5'),
])
def test_a_number_reads_the_way_its_format_says(code, value, shown):
    assert sheet.display_value(value, code) == shown


def test_currency_keeps_its_sign():
    assert sheet.display_value(1234.5, '#,##0.00\\ "€"') == '1.234,50 €'


def test_a_date_format_writes_a_date(tmp_path):
    import datetime

    assert sheet.display_value(datetime.date(2026, 12, 31), 'DD.MM.YYYY') == (
        '31.12.2026')


def test_a_format_this_grid_does_not_know_shows_the_plain_value():
    """A workbook can carry any format string. Guessing at what Excel would
    draw would put a number on screen that is not in the file."""
    assert sheet.display_value(1234.5, '[Red][>100]0.00;;') == '1234.5'


def test_text_is_never_reformatted():
    assert sheet.display_value('12.5', '#,##0.00') == '12.5'


# ---------------------------------------------------------------------------
# Reading what a cell looks like
# ---------------------------------------------------------------------------

@pytest.fixture
def book(tmp_path):
    workbook = openpyxl.Workbook()
    ws = workbook.active
    ws['A1'] = 'Kopf'
    ws['A1'].font = Font(bold=True, italic=True, underline='single')
    ws['A1'].fill = PatternFill('solid', fgColor='FFF2CC')
    ws['B1'] = 'Wert'
    ws['A2'] = 'x'
    ws['B2'] = 1234.5
    ws['B2'].number_format = '#,##0.00'
    path = tmp_path / 'tabelle.xlsx'
    workbook.save(path)
    return path


def test_a_cell_reports_only_what_differs_from_plain(book):
    """An empty dictionary per cell is what keeps the markup for a few
    thousand cells to a readable size."""
    _names, data = sheet.read_xlsx(book, None)
    assert data.styles[0][0] == {'b': 1, 'i': 1, 'u': 1, 'bg': 'FFF2CC'}
    assert data.styles[1][0] == {}


def test_a_number_format_is_applied_to_the_value_it_belongs_to(book):
    _names, data = sheet.read_xlsx(book, None)
    assert data.values[1][1] == '1.234,50'
    assert data.styles[1][1] == {'fmt': '#,##0.00'}


def test_white_is_not_a_colour_anyone_set(tmp_path):
    """Marking it would put a fill on every cell of a workbook that has
    none."""
    workbook = openpyxl.Workbook()
    workbook.active['A1'] = 'x'
    workbook.active['A1'].fill = PatternFill('solid', fgColor='FFFFFFFF')
    path = tmp_path / 'weiss.xlsx'
    workbook.save(path)
    _names, data = sheet.read_xlsx(path, None)
    assert 'bg' not in data.styles[0][0]


def test_a_colour_is_read_whatever_shape_it_arrives_in():
    assert sheet.check_fill('#fff2cc') == 'FFF2CC'
    assert sheet.check_fill('00FFF2CC') == 'FFF2CC'
    assert sheet.check_fill('') == ''
    with pytest.raises(sheet.SpreadsheetError):
        sheet.check_fill('bright red')


# ---------------------------------------------------------------------------
# The op
# ---------------------------------------------------------------------------

def test_a_format_op_is_checked_like_every_other():
    clean = sheet.validate_ops([
        {'op': 'format', 'r1': 1, 'c1': 1, 'r2': 2, 'c2': 3,
         'bold': True, 'fill': '#fff2cc'},
    ])
    assert clean == [{'op': 'format', 'r1': 1, 'c1': 1, 'r2': 2, 'c2': 3,
                      'bold': True, 'fill': 'FFF2CC'}]


def test_a_format_op_that_changes_nothing_is_refused():
    with pytest.raises(ValueError):
        sheet.validate_ops([{'op': 'format', 'r1': 1, 'c1': 1,
                             'r2': 1, 'c2': 1}])


def test_a_range_that_is_not_a_range_is_refused():
    for bad in ({'r1': 0, 'c1': 1, 'r2': 1, 'c2': 1},
                {'r1': 5, 'c1': 1, 'r2': 2, 'c2': 1},
                {'r1': 1, 'c1': 3, 'r2': 1, 'c2': 1}):
        with pytest.raises(ValueError):
            sheet.validate_ops([{'op': 'format', 'bold': True, **bad}])


def test_the_window_shows_the_format_before_it_is_saved(book):
    """A pending format has to look applied, the same way a pending edit
    shows its new value."""
    _names, data = sheet.read_xlsx(book, None)
    sheet.replay_ops(data, sheet.validate_ops([
        {'op': 'format', 'r1': 2, 'c1': 1, 'r2': 2, 'c2': 1, 'bold': True},
    ]))
    assert data.styles[1][0] == {'b': 1}


def test_taking_a_format_off_again_is_the_same_op(book):
    _names, data = sheet.read_xlsx(book, None)
    sheet.replay_ops(data, sheet.validate_ops([
        {'op': 'format', 'r1': 1, 'c1': 1, 'r2': 1, 'c2': 1,
         'bold': False, 'fill': ''},
    ]))
    assert 'b' not in data.styles[0][0]
    assert 'bg' not in data.styles[0][0]


# ---------------------------------------------------------------------------
# Writing it into the workbook
# ---------------------------------------------------------------------------

def test_a_format_reaches_the_file(book):
    sheet.apply_ops_xlsx(book, 'Sheet', [
        {'op': 'format', 'r1': 2, 'c1': 1, 'r2': 2, 'c2': 2,
         'bold': True, 'fill': 'DDEBF7', 'number_format': '0.00'},
    ], backup=False)

    reopened = openpyxl.load_workbook(book)
    cell = reopened.active['A2']
    assert cell.font.bold is True
    assert cell.fill.fgColor.rgb.endswith('DDEBF7')
    assert cell.number_format == '0.00'


def test_what_the_grid_does_not_show_is_not_thrown_away(tmp_path):
    """A workbook carries fonts, sizes and borders this grid never shows.
    Replacing the whole style for the sake of one bold would lose them."""
    workbook = openpyxl.Workbook()
    ws = workbook.active
    ws['A1'] = 'x'
    ws['A1'].font = Font(name='Georgia', size=14, italic=True)
    path = tmp_path / 'schrift.xlsx'
    workbook.save(path)

    sheet.apply_ops_xlsx(path, 'Sheet', [
        {'op': 'format', 'r1': 1, 'c1': 1, 'r2': 1, 'c2': 1, 'bold': True},
    ], backup=False)

    font = openpyxl.load_workbook(path).active['A1'].font
    assert font.bold is True
    assert font.name == 'Georgia' and font.size == 14
    assert font.italic is True, 'a change to one attribute cleared another'


def test_a_fill_can_be_taken_off(book):
    sheet.apply_ops_xlsx(book, 'Sheet', [
        {'op': 'format', 'r1': 1, 'c1': 1, 'r2': 1, 'c2': 1, 'fill': ''},
    ], backup=False)
    fill = openpyxl.load_workbook(book).active['A1'].fill
    assert fill.patternType in (None, 'none')


def test_a_csv_says_it_cannot_hold_formatting(tmp_path):
    """Dropping it quietly would leave the user with a table they think
    they coloured."""
    path = tmp_path / 'liste.csv'
    path.write_text('a;b\n1;2\n', encoding='utf-8')
    with pytest.raises(sheet.SpreadsheetError) as excinfo:
        sheet.apply_ops_delimited(
            path, [{'op': 'format', 'r1': 1, 'c1': 1, 'r2': 1, 'c2': 1,
                    'bold': True}], ';', backup=False)
    assert '.xlsx' in str(excinfo.value)


# ---------------------------------------------------------------------------
# The controls
# ---------------------------------------------------------------------------

def _grid(kind='xlsx'):
    data = sheet.SheetData(name='Sheet1', values=[['a']], total_rows=1,
                           total_cols=1, col_widths=[110])
    markup = sheet.render_grid_html(data, sheet_names=['Sheet1'], token='tok',
                                    kind=kind, editable=True)
    return markup[markup.index('</style>'):]


def test_the_controls_are_offered_on_a_workbook():
    markup = _grid()
    assert 'dsheet-b' in markup and 'dsheet-i' in markup and 'dsheet-u' in markup
    assert 'dsheet-swatch' in markup
    assert 'dsheet-numfmt' in markup


def test_a_csv_is_not_offered_them():
    assert 'dsheet-swatch' not in _grid(kind='csv')


def test_the_shortcuts_are_the_ones_people_already_use():
    script = sheet.grid_js('calc-scope-1', 'tok')
    for key in ("e.key === 'b'", "e.key === 'i'", "e.key === 'u'"):
        assert key in script


def test_the_cells_change_before_the_kernel_answers():
    """Waiting for a round trip would make every click feel like a request
    rather than like formatting."""
    script = sheet.grid_js('calc-scope-1', 'tok')
    body = script[script.index('function applyFormat'):][:900]
    assert 'paintCell' in body
    assert body.index('paintCell') < body.index('push(ops)') if 'push(ops)' in body else True


def test_one_undo_takes_the_formatting_off_again():
    script = sheet.grid_js('calc-scope-1', 'tok')
    body = script[script.index('function applyFormat'):][:1400]
    assert 'entry.before' in body, 'nothing captured what the cells looked like'
    assert 'remember({' in body


def test_each_cell_keeps_its_own_previous_look():
    """A range is rarely uniform -- half of it may already be bold -- so
    putting it back needs each cell's own state, not the range's."""
    script = sheet.grid_js('calc-scope-1', 'tok')
    assert 'function formatOps(cells, pick)' in script
    body = script[script.index('function formatOps'):][:500]
    assert 'r1: entry.row' in body and 'r2: entry.row' in body
