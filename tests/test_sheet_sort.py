"""Sorting that reorders the sheet, not just the view.

Clicking a heading has always sorted what is on screen and changed
nothing. That is the one to reach for by accident, so it stays as it is;
sorting the file is asked for by name in the right-click menu.

It runs over the whole sheet rather than the window: sorting the rows that
happen to be visible would interleave them with the ones that are not, and
the file would be scrambled in a way nobody could see.
"""

from __future__ import annotations

import json

import pytest

openpyxl = pytest.importorskip('openpyxl')
from openpyxl.styles import PatternFill

from delfin.dashboard import spreadsheet_view as sheet


@pytest.fixture
def book(tmp_path):
    workbook = openpyxl.Workbook()
    ws = workbook.active
    for row in (['Name', 'Betrag'], ['Zeta', 5], ['Alpha', 20],
                ['Mitte', None], ['Beta', 1]):
        ws.append(row)
    ws['A2'].fill = PatternFill('solid', fgColor='FFF2CC')
    path = tmp_path / 'buchungen.xlsx'
    workbook.save(path)
    return path


def _rows(path):
    return [row for row in openpyxl.load_workbook(path).active.iter_rows(
        values_only=True)]


# ---------------------------------------------------------------------------
# The order
# ---------------------------------------------------------------------------

def test_sorting_reorders_the_file(book):
    sheet.sort_sheet(book, 'Sheet', 1, backup=False)
    assert [row[0] for row in _rows(book)] == [
        'Name', 'Alpha', 'Beta', 'Mitte', 'Zeta']


def test_the_other_way_round(book):
    sheet.sort_sheet(book, 'Sheet', 1, descending=True, backup=False)
    assert [row[0] for row in _rows(book)] == [
        'Name', 'Zeta', 'Mitte', 'Beta', 'Alpha']


def test_numbers_sort_as_numbers_not_as_text(tmp_path):
    workbook = openpyxl.Workbook()
    for row in ([2], [10], [1]):
        workbook.active.append(row)
    path = tmp_path / 'zahlen.xlsx'
    workbook.save(path)
    sheet.sort_sheet(path, 'Sheet', 1, backup=False)
    assert [row[0] for row in _rows(path)] == [1, 2, 10]


def test_blank_cells_sink_whichever_way_the_sort_goes(book):
    """A column sorted descending should not start with its empty rows."""
    sheet.sort_sheet(book, 'Sheet', 2, backup=False)
    assert _rows(book)[-1][1] is None
    sheet.sort_sheet(book, 'Sheet', 2, descending=True, backup=False)
    assert _rows(book)[-1][1] is None


def test_the_whole_row_travels_with_the_cell_it_was_sorted_by(book):
    sheet.sort_sheet(book, 'Sheet', 2, backup=False)
    by_name = {row[0]: row[1] for row in _rows(book)[1:]}
    assert by_name == {'Beta': 1, 'Zeta': 5, 'Alpha': 20, 'Mitte': None}


# ---------------------------------------------------------------------------
# The heading
# ---------------------------------------------------------------------------

def test_a_heading_stays_at_the_top(book):
    sheet.sort_sheet(book, 'Sheet', 1, backup=False)
    assert _rows(book)[0] == ('Name', 'Betrag')


def test_the_heading_is_read_across_the_row_not_down_the_column():
    """A column of names under "Name" is text over text and says nothing;
    the "Betrag" beside it sits over numbers and settles it."""
    rows = [['Name', 'Betrag'], ['Zeta', 5], ['Alpha', 20]]
    assert sheet.looks_like_a_header(rows) is True


def test_data_with_no_heading_is_left_whole():
    assert sheet.looks_like_a_header([[3], [1], [2]]) is False


def test_with_nothing_to_go_on_the_first_row_is_taken_as_a_heading():
    """Which is what a spreadsheet assumes too. Sorting a heading into the
    middle of the data is the worse mistake, and not one that can be
    spotted by looking at the top of the sheet."""
    assert sheet.looks_like_a_header([['Ort'], ['Koeln'], ['Bonn']]) is True


def test_the_caller_can_say_so_instead(book):
    sheet.sort_sheet(book, 'Sheet', 1, header=False, backup=False)
    assert _rows(book)[0][0] == 'Alpha', 'the heading was kept anyway'


# ---------------------------------------------------------------------------
# What travels with a row
# ---------------------------------------------------------------------------

def test_a_cell_keeps_its_colour_when_its_row_moves(book):
    sheet.sort_sheet(book, 'Sheet', 1, backup=False)
    reopened = openpyxl.load_workbook(book).active
    coloured = [cell.row for cell in reopened['A']
                if cell.fill.patternType == 'solid']
    assert [reopened.cell(row=r, column=1).value for r in coloured] == ['Zeta']


def test_a_formula_still_points_at_its_own_row(tmp_path):
    """A reference that pointed at the row it sits in still does. Left
    alone, it would quietly read whichever row moved into its place."""
    workbook = openpyxl.Workbook()
    ws = workbook.active
    ws.append(['Name', 'Betrag', 'Doppelt'])
    for name, amount, row in (('Zeta', 5, 2), ('Alpha', 20, 3), ('Beta', 1, 4)):
        ws.append([name, amount, f'=B{row}*2'])
    path = tmp_path / 'formeln.xlsx'
    workbook.save(path)

    sheet.sort_sheet(path, 'Sheet', 1, backup=False)
    rows = _rows(path)
    assert [row[0] for row in rows] == ['Name', 'Alpha', 'Beta', 'Zeta']
    assert [row[2] for row in rows[1:]] == ['=B2*2', '=B3*2', '=B4*2']


# ---------------------------------------------------------------------------
# Care
# ---------------------------------------------------------------------------

def test_a_copy_is_kept_first(book, tmp_path):
    made, _kept = sheet.sort_sheet(book, 'Sheet', 1,
                                   backup_dir=book.parent / 'Backups')
    assert made is not None
    assert [row[0] for row in _rows(made)][1] == 'Zeta', (
        'the copy is of the sheet before it was sorted')


def test_a_sheet_already_in_order_is_not_rewritten(book):
    sheet.sort_sheet(book, 'Sheet', 1, backup=False)
    before = book.read_bytes()
    made, _kept = sheet.sort_sheet(book, 'Sheet', 1, backup=False)
    assert made is None
    assert book.read_bytes() == before


def test_a_sheet_that_is_not_there(book):
    with pytest.raises(sheet.SpreadsheetError):
        sheet.sort_sheet(book, 'Nope', 1, backup=False)


def test_a_column_that_is_not_there(book):
    with pytest.raises(sheet.SpreadsheetError):
        sheet.sort_sheet(book, 'Sheet', 0, backup=False)


def test_one_row_is_nothing_to_sort(tmp_path):
    workbook = openpyxl.Workbook()
    workbook.active.append(['nur eine'])
    path = tmp_path / 'eine.xlsx'
    workbook.save(path)
    assert sheet.sort_sheet(path, 'Sheet', 1, backup=False) == (None, False)


# ---------------------------------------------------------------------------
# How it is asked for
# ---------------------------------------------------------------------------

def test_it_is_asked_for_by_name_not_by_clicking_a_heading():
    """Clicking a heading still sorts the view and changes nothing. That is
    the one to reach for by accident."""
    script = sheet.grid_js('calc-scope-1', 'tok')
    assert "Sort A → Z (whole sheet)" in script
    assert "send('sort', [], {col: colNo})" in script
    body = script[script.index('function applySort'):][:900]
    assert "send('sort'" not in body, 'the heading click writes the file'


def test_a_csv_is_not_offered_it():
    script = sheet.grid_js('calc-scope-1', 'tok')
    body = script[script.index('Sort A → Z') - 300:script.index('Sort A → Z')]
    assert "dataset.kind === 'xlsx'" in body


# ---------------------------------------------------------------------------
# Through the tab
# ---------------------------------------------------------------------------

@pytest.fixture
def office(tmp_path, book):
    from delfin.dashboard import tab_calculations_browser as browser
    from delfin.dashboard.context import DashboardContext

    (tmp_path / 'archive').mkdir()
    ctx = DashboardContext(calc_dir=tmp_path, archive_dir=tmp_path / 'archive',
                           office_dir=tmp_path)
    ctx.run_js = lambda _s: None
    _widget, refs = browser.create_tab(ctx)
    refs['calc_list_directory']()
    file_list = refs['calc_file_list']
    match = [o for o in file_list.options if 'buchungen' in str(o)]
    value = match[0][1] if isinstance(match[0], tuple) else match[0]
    file_list.value = (value,) if isinstance(file_list.value, tuple) else value
    return book, refs


def _send(refs, action, **extra):
    state = refs['xyz_batch_state']
    message = {'action': action, 'token': state['sheet_view']['token'],
               'ops': [], 'cols': [], 'scroll': 0, 'cur': [0, 1]}
    message.update(extra)
    refs['calc_sheet_payload_input'].value = json.dumps(message)
    refs['calc_sheet_action_btn'].click()


def test_sorting_through_the_tab_writes_and_says_what_it_did(office):
    path, refs = office
    _send(refs, 'sort', col=1)
    assert [row[0] for row in _rows(path)][1] == 'Alpha'
    info = refs['calc_file_info'].value
    assert 'Sorted by column A' in info
    assert 'heading' in info
    assert 'backup' in info


def test_sorting_is_refused_while_edits_are_pending(office):
    path, refs = office
    before = _rows(path)
    _send(refs, 'edit', ops=[{'op': 'set', 'row': 2, 'col': 1, 'text': 'x'}])
    _send(refs, 'sort', col=1)
    assert _rows(path) == before
    assert 'Save or discard first' in refs['calc_file_info'].value
